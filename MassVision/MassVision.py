
"""
MassVision
module access within Slicer Python console:
	MassVision = slicer.modules.massvision.widgetRepresentation().self()
	peaks = MassVision.logic.peaks
"""

import vtk, qt, slicer
from slicer.ScriptedLoadableModule import *
from slicer.util import VTKObservationMixin
import logging
import os, re
import numpy as np

try:
	import pandas as pd
except ModuleNotFoundError:
	slicer.util.pip_install("pandas")
	import pandas as pd

# from MassVisionLib.Logic import * 


class MassVision(ScriptedLoadableModule):
	"""
	Uses ScriptedLoadableModule base class, available at:
	https://github.com/Slicer/Slicer/blob/master/Base/Python/slicer/ScriptedLoadableModule.py
	"""

	def __init__(self, parent):
		ScriptedLoadableModule.__init__(self, parent)

		self.parent.title = "MassVision"  
		self.parent.categories = ["Spectral Imaging"]
		self.parent.dependencies = []
		self.parent.contributors = ["Amoon Jamzad (Med-i Lab, Queen's University), Jade Warren, Ayesha Syeda)"] 
		self.parent.helpText = """
		MassVision is a software solution developed in 3D Slicer platform for end-to-end AI-driven analysis of Mass Spectrometry Imaging (MSI) data. 
		
		The functionalities include data exploration via various targeted, untargeted, and local-contrast visualization, co-localization with reference modality (histopathology annotations), dataset curation with spatial- and spectral-guidance, multi-slide dataset merge via feature alignment, spatial and spectral filtering, statistical analysis, feature ranking and selection, AI model training and validation, and whole-slide AI model deployment.		
		
		Please cite the following publication: Jamzad, A.; Warren, J.; Syeda, A.; Kaufmann, M.; Iaboni, N.; Nicol, C.; Rudan, J.; Ren, K.; Hurlbut, D.; Varma, S.; Fichtinger, G.; Mousavi, P. MassVision: An Open-Source End-to-End Platform for AI-Driven Mass Spectrometry Imaging Analysis. Analytical Chemistry 2025. https://doi.org/10.1021/acs.analchem.5c04018.
		"""
		self.parent.acknowledgementText = """

		"""

	
	def setupUi(self, parent):
		# Call the superclass implementation
		ScriptedLoadableModule.setupUi(self, parent)

		# Get a reference to your QTabWidget
		tabWidget = self.ui.tabWidget


class MassVisionTest(ScriptedLoadableModuleTest):
	"""
	This is the test case for your scripted module.
	Uses ScriptedLoadableModuleTest base class, available at:
	https://github.com/Slicer/Slicer/blob/master/Base/Python/slicer/ScriptedLoadableModule.py
	"""

	def setUp(self):
		""" Do whatever is needed to reset the state - typically a scene clear will be enough.
		"""
		slicer.mrmlScene.Clear()

	def runTest(self):
		"""Run as few or as many tests as needed here.
		"""
		self.setUp()
		self.test_MassVision1()

	def test_MassVision1(self):
		""" Ideally you should have several levels of tests.  At the lowest level
		tests should exercise the functionality of the logic with different inputs
		(both valid and invalid).  At higher levels your tests should emulate the
		way the user would interact with your code and confirm that it still works
		the way you intended.
		One of the most important features of the tests is that it should alert other
		developers when their changes will have an impact on the behavior of your
		module.  For example, if a developer removes a feature that you depend on,
		your test should break so they know that the feature is needed.
		"""

		self.delayDisplay("Starting the test")

		# logic = MassVisionLogic()

		self.delayDisplay('Test passed')

class MassVisionWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):
	"""
 	Uses ScriptedLoadableModuleWidget base class, available at:
	https://github.com/Slicer/Slicer/blob/master/Base/Python/slicer/ScriptedLoadableModule.py
	"""

	def __init__(self, parent=None):
		"""
		Called when the user opens the module the first time and the widget is initialized.
		"""
		ScriptedLoadableModuleWidget.__init__(self, parent)
		VTKObservationMixin.__init__(self)  # needed for parameter node observation
		self.logic = None
		self._parameterNode = None
		self._updatingGUIFromParameterNode = False
		self.parameterSetNode = None
		self.REIMS = 0
		self.AppMode = 0 #0:MSI, 1:Embeddings
		self.EmbedColor = "#A35C36"
		self.EmbedColor_btn = "#72300C" #80360E
		self._blendCtrl = None

	def setup(self):
		"""
		Called when the user opens the module the first time and the widget is initialized.
		"""
		ScriptedLoadableModuleWidget.setup(self)

		from MassVisionLib.Logic import MassVisionLogic

		# Load widget from .ui file (created by Qt Designer).
		uiWidget = slicer.util.loadUI(self.resourcePath('UI/MassVision.ui'))
		self.layout.addWidget(uiWidget)
		self.ui = slicer.util.childWidgetVariables(uiWidget)

		self.cases_config = {}
		self.logic = MassVisionLogic()
		self.logic.AppMode = self.AppMode

		# Set scene in MRML widgets. Make sure that in Qt designer the top-level qMRMLWidget's
		# "mrmlSceneChanged(vtkMRMLScene*)" signal in is connected to each MRML widget's.
		# "setMRMLScene(vtkMRMLScene*)" slot.
		#self.ui.subjectHierarchy.setMRMLScene(slicer.mrmlScene)

		# Create logic class. Logic implements all computations that should be possible to run
		# in batch mode, without a graphical user interface.


		

		# set the first tab as the default loading tab
		self.ui.tabWidget.setCurrentIndex(0)
		self.ui.tabWidget.currentChanged.connect(self.onTabChange)
		slicer.util.moduleSelector().connect('moduleSelected(QString)', self.onModuleChange)
		
		# make markup toolbar visible
		slicer.util.mainWindow().findChild('QToolBar', 'MarkupsToolBar').show()
		
		# Connections
		if self.REIMS==0:
			self.ui.label_REIMS.hide()
			self.ui.REIMSFileSelect.hide()
			self.ui.filenameREIMSBrowser.hide()
			self.ui.REIMSFileLoad.hide()
			self.ui.REIMSInformation.hide()

		self.ui.LoadingsInfo.hide()
		self.ui.pcaExtendCheckbox.hide()
		self.ui.roiCintrastExtend.hide()
		# self.ClearClusterTable()
		## hide the in-ui table view
		# self.ui.ClusterTable.setVisible(False)

		self.ui.csvLoad.hide()
		self.ui.modellingFile.hide()
		self.ui.deployImport.hide()
		self.ui.deployModelImport.hide()

		norm_methods = ['Total ion current (TIC)', 'Total signal current (TSC)', 'Root mean square (RMS)', 'Median', 'Mean', 'Reference ion']
		for method in norm_methods:
			self.ui.normMethodComboBox.addItem(method)
		self.ui.normMethodComboBox.setCurrentText("Total ion current (TIC)")
		self.ui.refionLabel.setVisible(False)
		self.ui.refIoncomboBox.setVisible(False)
		self.ui.thresholdLabel.setVisible(False)
		self.ui.thresholdValue.setVisible(False)

		for method in norm_methods:
			self.ui.depNormMethod.addItem(method)
		self.ui.depNormMethod.setCurrentText("Total ion current (TIC)")
		self.ui.depRefIonLab.setVisible(False)
		self.ui.depComboboxIon.setVisible(False)
		self.ui.depNormThreshLab.setVisible(False)
		self.ui.depNormThresh.setVisible(False)

		self.ui.MLlabel2.setVisible(False)
		self.ui.MLparam2.setVisible(False)

		self.ui.AlignTolLabel.setVisible(False)
		self.ui.AlignTolVal.setVisible(False)
		self.ui.AlignBinLab.setVisible(False)
		self.ui.AlignBinVal.setVisible(False)
		
		self.ui.lockmassLab.setVisible(False)
		self.ui.lockmassVal.setVisible(False)
		self.ui.rawrangeStLab.setVisible(False)
		self.ui.rawrangeStVal.setVisible(False)
		self.ui.rawrangeEnLab.setVisible(False)
		self.ui.rawrangeEnVal.setVisible(False)
		self.ui.rawsmoothLab.setVisible(False)
		self.ui.rawsmoothVal.setVisible(False)

		self.ui.statGroup1Lab.setVisible(False)
		self.ui.statGroup1combo.setVisible(False)
		self.ui.statGroup2Lab.setVisible(False)
		self.ui.statGroup2combo.setVisible(False)


		# Set logo in UI
		logo_path = self.resourcePath('Icons/UI_nameM.png')
		# logo_path = self.resourcePath('Icons/UI_logoS.png')
		self.ui.logo.setPixmap(qt.QPixmap(logo_path))

		# Set pushbutton icons
		icon_path = self.resourcePath('Icons/roi.png')
		self.ui.ROIforLocalContrast.setIcon(qt.QIcon(icon_path))

		icon_path = self.resourcePath('Icons/marker.png')
		self.ui.RAWplaceFiducial.setIcon(qt.QIcon(icon_path))
		self.ui.placeFiducial.setIcon(qt.QIcon(icon_path))
		self.ui.placeFiducial_sim.setIcon(qt.QIcon(icon_path))

		# Set tab widget tooltip and icons
		# ---Robert Added Icon Name 'label'------
		icon_names = ['home', 'file', 'visualization', 'dataset', 'alignment', 'preprocess', 'stat', 'train', 'report', 'inference', 'pathway']
		for i in range(self.ui.tabWidget.count):
			tabText = self.ui.tabWidget.tabText(i)
			self.ui.tabWidget.setTabText(i, "")            
			self.ui.tabWidget.tabBar().setTabToolTip(i, tabText)  
			icon_path = self.resourcePath(f'Icons/{icon_names[i]}.png')
			self.ui.tabWidget.setTabIcon(i, qt.QIcon(icon_path))

		self.ui.tabWidget.setTabText(0, 'Home') 
		self.ui.tabWidget.tabBar().setIconSize(qt.QSize(25, 25))

		# Collapse the Data Probe
		dataProbeWidget = slicer.util.mainWindow().findChild(qt.QWidget, "DataProbeCollapsibleWidget")
		if dataProbeWidget and hasattr(dataProbeWidget, "collapsed"):
			dataProbeWidget.collapsed = True

		# Hide additional toolbars
		for child in slicer.util.mainWindow().findChildren(qt.QToolBar):
			visible_toolbars = ['ModuleSelectorToolBar', 'ViewToolBar', 'MouseModeToolBar', 'DialogToolBar']
			if child.objectName in visible_toolbars:
				child.setVisible(True)
			else:
				child.setVisible(False)

		# Collapse Python console
		pythonConsoleDock = slicer.util.mainWindow().findChild(qt.QDockWidget, "PythonConsoleDockWidget")
		if pythonConsoleDock:
			pythonConsoleDock.setVisible(False)

		# Layout to red view
		slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUpRedSliceView)

		# Heatmap list singleIonHeatmapList
		self.ui.singleIonHeatmapList.addItem('Inferno')
		self.ui.singleIonHeatmapList.addItem('DivergingBlueRed')
		self.ui.singleIonHeatmapList.addItem('PET-Rainbow2')
		self.ui.singleIonHeatmapList.addItem('Cividis')
		self.ui.singleIonHeatmapList.addItem('ColdToHotRainbow')

		for i in range(self.ui.singleIonHeatmapList.count):
			text = self.ui.singleIonHeatmapList.itemText(i)
			data = self.ui.singleIonHeatmapList.itemData(i)
			self.ui.RawImgHeatmap.addItem(text, data)

		# Home tab
		self.ui.clearReloadPush.connect("clicked(bool)", self.onClearReload)
		self.ui.loadScenePush.connect("clicked(bool)", self.onLoadScene)

		for i in range(1, 11):
			button = getattr(self.ui, f"Go2tab{i}")
			button.clicked.connect(lambda checked=False, idx=i: self.ui.tabWidget.setCurrentIndex(idx))

		self.ui.userManual.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://slicermassvision.readthedocs.io/")))
		self.ui.sampleData.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://github.com/jamzad/SlicerMassVision/releases/tag/test-data")))
		self.ui.codeBase.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://github.com/jamzad/SlicerMassVision")))
		self.ui.publication.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://pubs.acs.org/doi/10.1021/acs.analchem.5c04018")))

		self.ui.database1.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://hmdb.ca/")))
		self.ui.database2.clicked.connect(
			lambda: qt.QDesktopServices.openUrl(qt.QUrl("https://www.lipidmaps.org/")))

		# Data Import
		self.ui.textFileSelect.connect("clicked(bool)", self.onTextFileSelect)

		self.ui.histoFileSelect.connect("clicked(bool)",self.onHistoSelect)

		self.ui.rawSelect.connect("clicked(bool)", self.onRawSelect)
		self.ui.RAWplaceFiducial.connect("clicked(bool)", lambda checked: self.onPutFiducial("raw-spectrum"))
		self.ui.RAWplotSpectra.connect("clicked(bool)", self.onRawPlotSpectra)
		self.ui.RawPlotImg.connect("clicked(bool)", self.onRawPlotImg)

		self.logic.raw_image_tol = float(self.ui.RawImgTol.text)
		self.ui.RawImgTol.textChanged.connect(self.onRawImgTolChange)
		
		self.ui.rawsmoothCheck.connect("clicked(bool)", self.onRawsmoothCheck)
		self.ui.lockmassCheck.connect("clicked(bool)", self.onLockmassCheck)
		self.ui.rawrangeCheck.connect("clicked(bool)", self.onRawrangCheck)
		self.ui.rawProcess.connect("clicked(bool)", self.onRawProcess)
		
		# ----- Robert button connection --------
		self.ui.labelpeaksbutton.connect("clicked(bool)", self.onLabelPeaks)
		self.ui.HMDBDownloadpushButton.connect("clicked(bool)", self.onUpdateHMDBDatabase)

		self.ui.ExportPushBotton.connect("clicked(bool)",self.onExport)

		self.ui.REIMSFileSelect.connect("clicked(bool)",self.onREIMSSelect)
		self.ui.REIMSFileLoad.connect("clicked(bool)",self.onREIMSLoad)


		# Visualization
		self.ui.visNorm_spectra.currentTextChanged.connect(self.visRenormalize)
		self.ui.visNorm_ions.currentTextChanged.connect(self.visRenormalize)

		self.ui.spectrumPlot.connect("clicked(bool)", self.onSpectrumPlot)
		self.ui.placeFiducial.connect("clicked(bool)", lambda checked: self.onPutFiducial( ["spectrum", "pixel"][self.AppMode] ))

		self.ui.simHeatmap.connect("clicked(bool)", self.onSimHeatmap)
		self.ui.SimThumbnail.connect("clicked(bool)", self.onSimThumbnail)
		self.ui.placeFiducial_sim.connect("clicked(bool)", lambda checked: self.onPutFiducial("similarity", single_point=True))

		self.ui.AbundanceThumbnail.connect("clicked(bool)", self.onAbundanceThumbnail)

		self.ui.singleIonButton.connect("clicked(bool)", self.selectedSingleIon)
		self.ui.multiIonButton.connect("clicked(bool)", self.selectedMultiIon)
		self.ui.PCA_button.connect("clicked(bool)", self.onPCAButton)
		self.ui.partialPCA.connect("clicked(bool)", self.onPartialPCAButton)
		self.ui.ROIforLocalContrast.connect("clicked(bool)", self.onROIforLocalContrast)
		self.dataInfo = ''
		self.ui.ContrastThumbnail.connect("clicked(bool)", self.onContrastThumbnail)

		self.ui.saveProjection.connect("clicked(bool)", self.onSaveProjection)
		self.ui.loadProjection.connect("clicked(bool)", self.onLoadProjection)

		self.ui.NLVisMethod.currentTextChanged.connect(self.onNLVisMethod)
		self.ui.UmapButton.connect("clicked(bool)", self.onUMAPVis)

		self.ui.Cluster_button.connect("clicked(bool)", self.onClusterButton)
		self.ui.ClusterThumbnail.connect("clicked(bool)", self.onClusterThumbnail)

		#### Blending
		# 1) Node selectors must know the scene
		self.ui.bgVolumeSelector.setMRMLScene(slicer.mrmlScene)
		self.ui.fgVolumeSelector.setMRMLScene(slicer.mrmlScene)

		# 2) Create controller handle (None until enabled)
		self.ui.blendGroupBox.checked = False
		self._setBlendWidgetsEnabled(False)
		self._blendCtrl = None

		# 3) Debounce timer for selector changes (prevents double-firing)
		self._blendDebounce = qt.QTimer()
		self._blendDebounce.setSingleShot(True)
		self._blendDebounce.timeout.connect(self._applyBlendInputs)

		# 4) Connect UI signals
		self.ui.blendGroupBox.toggled.connect(self._onBlendEnabledChanged)

		# qMRMLNodeComboBox commonly emits currentNodeChanged(vtkMRMLNode*)
		self.ui.bgVolumeSelector.currentNodeChanged.connect(lambda _n: self._scheduleBlendApply())
		self.ui.fgVolumeSelector.currentNodeChanged.connect(lambda _n: self._scheduleBlendApply())

		self.ui.overlayRadioButton.toggled.connect(self._onBlendModeToggled)
		self.ui.wipeRadioButton.toggled.connect(self._onBlendModeToggled)
		self.ui.verticalWipeRadioButton.toggled.connect(self._onBlendModeToggled)

		self.ui.blendSlider.valueChanged.connect(self._onBlendSliderChanged)

		self._setupBlendSliderSweepUI()
		self.ui.overlayRadioButton.checked = True
		#### Blending

		# Dataset generation
		self.ui.gotoRegistration.connect("clicked(bool)", self.landmark)

		self.ui.segVolCombo1.setMRMLScene(slicer.mrmlScene)
		self.ui.segVolCombo2.setMRMLScene(slicer.mrmlScene)

		self.ui.segmentEditor.connect("clicked(bool)", self.showSegmentEditor)
		self.ui.roiContrast.connect("clicked(bool)", self.onROIContrast)
		self.ui.roiContrastLDA.connect("clicked(bool)", self.onROIContrastLDA)

		self.ui.roiSimilarity.connect("clicked(bool)", self.onROISimilarity)
		self.ui.roiExpand.connect("clicked(bool)", self.onROIExpand)

		self.ui.segmentVisibility.connect("clicked(bool)", self.onSegmentVisibility)

		self.ui.createCSVbutton.connect("clicked(bool)",self.onCSVconnect)
		self.ui.createMetadata.connect("clicked(bool)", lambda checked: self.onCSVconnect(meta_only=True))

		self.ui.saveScenePush.connect("clicked(bool)",self.onSaveScene)

		# Multi-slide alignment
		self.files = set()
		self.ui.selectToAlign.connect("clicked(bool)", self.onSelectToAlign)
		self.ui.loadToAlign.connect("clicked(bool)", self.onLoadToAlign)

		self.ui.alignPreview.connect("clicked(bool)", self.onAlignPreview)
		
		self.ui.AlignMatchMethod.currentTextChanged.connect(self.onAlignMatchMethod)

		self.ui.alignButton.connect("clicked(bool)", self.onMerge)
		
		# Dataset post-processing
		self.ui.csvSelect.connect("clicked(bool)", self.onCsvSelect)
		self.ui.csvLoad.connect("clicked(bool)", self.onCsvLoad)
		self.ui.normalizeCheckbox.connect("clicked(bool)", self.onNormalizationState)
		
		self.ui.normMethodComboBox.currentTextChanged.connect(self.onNormMethodChange)

		self.ui.spectrumFiltercheckBox.connect("clicked(bool)", self.onFilterState)

		self.ui.lowIntFiltercheckBox.connect("clicked(bool)", self.onIntFilterState)
		self.ui.lowVarFiltercheckBox.connect("clicked(bool)", self.onVarFilterState)

		self.ui.pixelaggcheckBox.connect("clicked(bool)", self.onAggState)
		self.ui.applyProcessingButton.connect("clicked(bool)", self.onApplyProcessing)	
  

		# Statistical analysis
		self.ui.importStat.connect("clicked(bool)", self.onSelectStatData)
		self.ui.distributionPCA.connect("clicked(bool)", self.onPlotDIstribution)
		self.ui.boxplotButton.connect("clicked(bool)", self.onBoxPlot)
		self.ui.anovaButton.connect("clicked(bool)", self.onANOVA)
		self.ui.statClassConfig.currentTextChanged.connect(self.onStatConfigChange)
		self.ui.ttestButton.connect("clicked(bool)", self.onTtest)
		self.ui.volcanoButton.connect("clicked(bool)", self.onVolcano)

		
		# Model Training
		self.ui.selectCSV.connect("clicked(bool)", self.onSelectModelData)
		self.ui.modellingFile.connect("clicked(bool)", self.onModellingLoad)

		self.ui.FRankMethod.currentTextChanged.connect(self.onRankMethodChange)
		self.ui.FRankApply.clicked.connect(self.onFeatureRank)
		
		self.ui.FSelManualUpload.setVisible(False)
		self.ui.FnumberLabel.setVisible(False)
		self.ui.FnumberValue.setVisible(False)
		self.ui.FSelMethod.currentTextChanged.connect(self.onSelMethodChange)
		self.ui.FSelManualUpload.clicked.connect(self.onFeatureListUpload)

		self.ui.randomSplit.connect("clicked(bool)", self.onRandomSplit) # random
		self.ui.customSplit.connect("clicked(bool)", self.onCustomSplit)  # custom
		self.ui.allTrain.connect("clicked(bool)", self.onAllTrainSplit)  # all train
		self.ui.XVall.connect("clicked(bool)", self.onCrossVal)  # all train

		self.ui.trainModel.connect("clicked(bool)", self.onModelTrain)
		self.onAllTrainSplit()
		
		self.ui.ModelSelectCombobox.currentTextChanged.connect(self.onMLMethod)

		# Results
		self.model_results = ''

		# Deployment
		self.ui.deploySelect.connect("clicked(bool)", self.onDeploySelect)
		self.ui.deployImport.connect("clicked(bool)", self.onDeployLoad)
		self.ui.deployModelSel.connect("clicked(bool)", self.onDeployModelSel)
		self.ui.deployModelImport.connect("clicked(bool)", self.onDeployModelLoad)
		self.ui.deployNormcheck.connect("clicked(bool)", self.onDeployNormCheck)

		# self.ui.depRadioTIC.toggled.connect(self.onDepNormRadioToggle)
		# self.ui.depRadioIon.toggled.connect(self.onDepNormRadioToggle)

		self.ui.depNormMethod.currentTextChanged.connect(self.onDepNormMethodChange)

		self.ui.deployAGGcheck.connect("clicked(bool)", self.onDeployAggCheck)

		self.ui.depMaskcheck.connect("clicked(bool)", self.onDepMaskcheck)

		self.ui.depPCAVis.connect("clicked(bool)", self.onPCAButton)
		self.ui.depGoVisButton.connect("clicked(bool)", self.onDepGoVis)

		self.ui.depVisCombo.setMRMLScene(slicer.mrmlScene)
		self.ui.depVisCombo.setEnabled(False)
		
		self.ui.depGoSegEdButton.connect("clicked(bool)", self.onDepGoSeg)

		self.ui.deployRun.connect("clicked(bool)", self.onApplyDeployment)	

		# Pathway analysis

		# --- Robert Addition of Buttons and Items For Peak Labeling ---
		self.current_results_df = None
		self.ui.inputtedpeakslineedit.setPlaceholderText("e.g., 302.1594, 281.231")
		self.ui.moleculetoleranacelineedit.setPlaceholderText("e.g., 0.1, 0.005") 
		# Adduct button setup
		self.ui.exportpeaklabelsCSVbutton.connect('clicked(bool)', self.onExportPeakLabelExcel)
		self.ui.loadmzvaluescsvpushButton.connect('clicked(bool)', self.onLoadMzValuesCsv)
		# Radiobutton setup
		self.ui.findclosestcandidateradioButton.setChecked(True) # Set the default starting button
		self.buttonGroup = qt.QButtonGroup()
		self.buttonGroup.addButton(self.ui.findclosestcandidateradioButton)
		self.buttonGroup.addButton(self.ui.findallcandidatesradioButton)
		self.buttonGroup.buttonClicked.connect(self.onRadioButtonClicked)
		
		# --- Robert Addition for link opening ----
		self.ui.displaypatwaystextbrowser.setOpenLinks(False)
		self.ui.displaypatwaystextbrowser.anchorClicked.connect(self.onLinkClicked)

		# Create and setup browser view
		self._setupBrowserOnlyView()

		# --- UI Setup for Peak Labeling ---
			# Configure the display table created in Qt Designer
		self.ui.moleculesTableWidget.setColumnCount(7)
		self.ui.moleculesTableWidget.setHorizontalHeaderLabels(['Select', 'Searched m/z', 'Adduct', 'Molecule', 'Source ID','KEGG ID', 'Error'])
		header = self.ui.moleculesTableWidget.horizontalHeader()
			# Shrink the Checkbox, m/z, and Adduct columns to be as small as possible
		header.setSectionResizeMode(0, qt.QHeaderView.ResizeToContents) 
		header.setSectionResizeMode(1, qt.QHeaderView.ResizeToContents) 
		header.setSectionResizeMode(2, qt.QHeaderView.ResizeToContents)
			# Stretch the Molecule Name column to absorb all the extra empty space
		header.setSectionResizeMode(3, qt.QHeaderView.Stretch)
			# Shrink the KEGG ID and Error columns
		header.setSectionResizeMode(4, qt.QHeaderView.ResizeToContents)
		header.setSectionResizeMode(5, qt.QHeaderView.ResizeToContents)
		header.setSectionResizeMode(6, qt.QHeaderView.ResizeToContents)
			# Connect the button created in Qt Designer
		self.ui.searchPathwaysButton.connect('clicked(bool)', self.onSearchPathways)
		# ------- End of Robert Additions for this section --------


		# Mode change for ViT Embeddings
		modeButtonGroup = qt.QButtonGroup()
		modeButtonGroup.setExclusive(True)
		modeButtonGroup.addButton(self.ui.MSI_mode, 0)
		modeButtonGroup.addButton(self.ui.EMB_mode, 1)
		self.ui.modeButtonGroup = modeButtonGroup
		self.ui.modeButtonGroup.connect("buttonToggled(QAbstractButton*,bool)", self.onModeChangeButton)
		
		last_AppMode = slicer.app.settings().value("MassVision/Mode")
		if last_AppMode==1:
			self.ui.EMB_mode.setChecked(True)
			self.AppMode = 1

		# Make sure parameter node is initialized (needed for module reload)
		self.initializeParameterNode()


	def _setupBrowserOnlyView(self):
		#
		# Main browser widget
		#
		self.internalBrowser = slicer.qSlicerWebWidget()
		self.internalBrowser.handleExternalUrlWithDesktopService = False

		webView = self.internalBrowser.webView()

		#
		# Toolbar
		#
		self.browserToolbar = qt.QToolBar()
		self.browserToolbar.setMovable(False)
		self.browserToolbar.setFloatable(False)
		self.browserToolbar.setIconSize(qt.QSize(16, 16))
		self.browserToolbar.setStyleSheet("""
		QToolBar {
			background: #006666;
			border: 0px;
			spacing: 4px;
			padding: 2px;
		}
		QToolButton {
			padding: 4px 6px;
		}
		QLabel {
			font-weight: bold;
			padding-left: 4px;
			padding-right: 8px;
		}
		""")

		self.browserTitleLabel = qt.QLabel("Browser View ")
		self.browserToolbar.addWidget(self.browserTitleLabel)

		self.backAction = self.browserToolbar.addAction("Back")
		self.forwardAction = self.browserToolbar.addAction("Forward")
		self.reloadAction = self.browserToolbar.addAction("Reload")

		self.backAction.connect("triggered()", webView.back)
		self.forwardAction.connect("triggered()", webView.forward)
		self.reloadAction.connect("triggered()", webView.reload)

		#
		# Address bar
		#
		self.addressBar = qt.QLineEdit()
		self.addressBar.setPlaceholderText("Enter URL and press Enter")
		self.addressBar.connect("returnPressed()", self.onAddressEntered)
		self.browserToolbar.addWidget(self.addressBar)

		#
		# Keep toolbar state and address bar in sync
		#
		def updateNavigationState(*args):
			try:
				self.backAction.enabled = webView.history().canGoBack()
				self.forwardAction.enabled = webView.history().canGoForward()
			except Exception:
				self.backAction.enabled = True
				self.forwardAction.enabled = True

			try:
				self.addressBar.setText(webView.url().toString())
			except Exception:
				pass

		webView.loadFinished.connect(updateNavigationState)
		webView.urlChanged.connect(updateNavigationState)

		#
		# Keep toolbar state and address bar in sync
		#
		def updateNavigationState(*args):
			try:
				self.backAction.enabled = webView.history().canGoBack()
				self.forwardAction.enabled = webView.history().canGoForward()
			except Exception:
				self.backAction.enabled = True
				self.forwardAction.enabled = True

			try:
				self.addressBar.setText(webView.url().toString())
			except Exception:
				pass

		webView.loadFinished.connect(updateNavigationState)
		webView.urlChanged.connect(updateNavigationState)

		#
		# Container widget shown in the Slicer view area
		#
		self.browserViewWidget = qt.QWidget()
		browserLayout = qt.QVBoxLayout(self.browserViewWidget)
		browserLayout.setContentsMargins(0, 0, 0, 0)
		browserLayout.setSpacing(0)
		browserLayout.addWidget(self.browserToolbar)
		browserLayout.addWidget(self.internalBrowser)

		#
		# Register custom singleton view
		#
		self.browserViewFactory = slicer.qSlicerSingletonViewFactory()
		self.browserViewFactory.setTagName("MassVisionBrowserView")
		self.browserViewFactory.setWidget(self.browserViewWidget)
		slicer.app.layoutManager().registerViewFactory(self.browserViewFactory)

		#
		# Layout that contains only the browser view
		#
		self.browserOnlyLayoutId = 501

		layoutXml = """
		<layout type="vertical">
			<item>
			<MassVisionBrowserView />
			</item>
		</layout>
		"""

		layoutNode = slicer.app.layoutManager().layoutLogic().GetLayoutNode()
		layoutNode.AddLayoutDescription(self.browserOnlyLayoutId, layoutXml)

	def showBrowserOnlyView(self, url="https://example.com"):
		self.internalBrowser.setUrl(url)
		slicer.app.layoutManager().setLayout(self.browserOnlyLayoutId)

	def onAddressEntered(self):
		text = self.addressBar.text.strip()
		if not text:
			return

		# Add scheme if missing
		if "://" not in text:
			text = "https://" + text

		self.internalBrowser.setUrl(text)



	### Mode Selector
	def onModeChangeButton(self, btn, checked):
		if not checked:
			return
			
		selectedId = self.ui.modeButtonGroup.id(btn)
		self.AppMode = selectedId
		self.logic.AppMode = self.AppMode
		slicer.app.settings().setValue("MassVision/Mode", self.AppMode)

		if selectedId==0:
			slicer.mrmlScene.Clear()
			slicer.util.reloadScriptedModule('MassVision')
			self.resetTabwidgetScroll()
			print("Mode changed to MassVision")

		elif selectedId==1:
			slicer.mrmlScene.Clear()
			
			# disable irrelevant tabs
			# [self.ui.tabWidget.setTabEnabled(x,False) for x in range(4,self.ui.tabWidget.count)]
			[self.ui.tabWidget.setTabEnabled(x,False) for x in [4,9,10]]

			# change style and color
			target_bg  = "background-color: #006666;"
			target_fg  = "color: rgb(255, 255, 255);"
			newStyle = f"""
			QPushButton:enabled {{
				background-color: {self.EmbedColor_btn};
				color: rgb(255, 255, 255);
			}}
			"""
			for _, w in vars(self.ui).items():
				if isinstance(w, qt.QPushButton):
					ss = w.styleSheet
					if target_bg in ss and target_fg in ss:
						w.setStyleSheet(newStyle)
			
			btns = ["ROIforLocalContrast", "RAWplaceFiducial", "placeFiducial", "placeFiducial_sim"]
			[recolorButtonIcon(getattr(self.ui, btn), color=self.EmbedColor) for btn in btns]
			[recolorTabIcon(self.ui.tabWidget, i, color=self.EmbedColor) for i in range(self.ui.tabWidget.count)]

			# disable unnecessary buttons
			# for i in range(4, 10):  # 3 to 9 inclusive
			for i in [4,9,10]:
				btn = getattr(self.ui, f"Go2tab{i}")
				btn.setEnabled(False)
			# btns = ["database1", "database2"]
			# [getattr(self.ui, btn).setEnabled(False) for btn in btns ]

			# hide unnecessary actions
			objs = ["database1", "database2", "label_30",
				"CollapsibleRaw", "label_export", "ExportPushBotton",
				"label_68", "gotoRegistration", "hspacer_11", "vspacer_12"]
			[getattr(self.ui, obj).setVisible(False) for obj in objs ]

			# change labels
			updateUITexts(self.ui)
			self.ui.PCA_button.setText("PCA (global contrast)")
			self.ui.partialPCA.setText("P3CA (local contrast)")
			
			# Set logo in UI
			logo_path = self.resourcePath('Icons/embeddings_logo.png')
			self.ui.logo.setPixmap(qt.QPixmap(logo_path))

			# reset the scroll
			self.resetTabwidgetScroll()

			# change settings
			self.ui.visNorm_spectra.setCurrentIndex(0)

			self.ui.CollapsibleMode.collapsed = True
			print("Mode changed to EmbedVision")


	def resetTabwidgetScroll(self):
		scrollArea = self.ui.tabWidget.parent()
		while scrollArea and not isinstance(scrollArea, qt.QScrollArea):
			scrollArea = scrollArea.parent()
		if scrollArea:
			scrollArea.verticalScrollBar().setValue(0)

	### Data Import Tab

	def onClearReload(self):
		confirmation_dialog = qt.QMessageBox()
		confirmation_dialog.setIcon(qt.QMessageBox.Question)
		confirmation_dialog.setText("Are you sure you want to delete all files and clear the scene?")
		confirmation_dialog.setWindowTitle("Confirmation")
		confirmation_dialog.setStandardButtons(qt.QMessageBox.Yes | qt.QMessageBox.No)

		# Show the confirmation dialog and get the user's choice
		choice = confirmation_dialog.exec_()

		# Check the user's choice
		if choice == qt.QMessageBox.Yes:
			slicer.mrmlScene.Clear()
			slicer.util.reloadScriptedModule('MassVision')
			print('MODULE RELOADED')


	def onLoadScene(self):
		fileExplorer = qt.QFileDialog()
		mrbFilename = fileExplorer.getOpenFileName(None, "Load Slicer Project", "", "MSI Project Files (*.mrb);;All Files (*)")
		slicer.util.loadScene(mrbFilename)

	def onRawSelect(self):
		fileExplorer = qt.QFileDialog()
		filePath = fileExplorer.getOpenFileName(None, "Import raw MSI data", "", "imzML (*.imzml);;All Files (*)")
		if filePath:
			self.ui.rawLineEdit.setText(filePath)
			info, mz_range = self.logic.RawFileLoad(filePath)
			if info:
				self.ui.rawInfo.setText(info)
				self.ui.rawrangeStVal.setText( np.round(mz_range[0]) )
				self.ui.rawrangeEnVal.setText( np.round(mz_range[1]) )
				self.logic.saveFolder = os.path.dirname(filePath)
				self.logic.slideName = os.path.basename(filePath)

	def onPutFiducial(self, listName, single_point=False):
		scene = slicer.mrmlScene
		appLogic = slicer.app.applicationLogic()
		selectionNode = appLogic.GetSelectionNode()
		interactionNode = appLogic.GetInteractionNode()
		markupsLogic = slicer.modules.markups.logic()

		# Exact-name lookup only
		fiducialNode = slicer.util.getFirstNodeByClassByName(
			"vtkMRMLMarkupsFiducialNode", listName
		)

		# Recreate one-shot lists such as "similarity"
		if single_point and fiducialNode:
			scene.RemoveNode(fiducialNode)
			fiducialNode = None

		# Create node using Slicer's node factory
		if not fiducialNode:
			fiducialNode = scene.AddNewNodeByClass("vtkMRMLMarkupsFiducialNode", listName)
			fiducialNode.CreateDefaultDisplayNodes()

		# For single-point nodes, remove automatic numbering like "similarity_1"
		if single_point:
			observerTag = None

			def onPointAdded(caller, event):
				nonlocal observerTag
				if caller.GetNumberOfControlPoints() > 0:
					caller.SetNthControlPointLabel(0, listName)
					if observerTag is not None:
						caller.RemoveObserver(observerTag)
						observerTag = None

			observerTag = fiducialNode.AddObserver(
				slicer.vtkMRMLMarkupsNode.PointPositionDefinedEvent,
				onPointAdded
			)

		# Make this exact node the active placement target
		selectionNode.SetReferenceActivePlaceNodeClassName("vtkMRMLMarkupsFiducialNode")
		selectionNode.SetActivePlaceNodeID(fiducialNode.GetID())

		# Keep markups logic in sync too
		markupsLogic.SetActiveListID(fiducialNode)

		# Enter single-place mode
		interactionNode.SetPlaceModePersistence(0)
		interactionNode.SetCurrentInteractionMode(interactionNode.Place)

		return fiducialNode

	def onRawPlotSpectra(self):
		self.logic.RawPlotSpectra()

	def onRawPlotImg(self):
		ion_mz = float(self.ui.RawImgIon.text)
		tol_mz = float(self.ui.RawImgTol.text)
		img_heatmap = self.ui.RawImgHeatmap.currentText
		self.logic.RawPlotImg(ion_mz, tol_mz, img_heatmap)
	
	def onRawImgTolChange(self, text):
		self.logic.raw_image_tol = float(text)

	def onRawsmoothCheck(self):
		currentState = self.ui.rawsmoothCheck.isChecked()
		self.ui.rawsmoothLab.setVisible(currentState)
		self.ui.rawsmoothVal.setVisible(currentState)

	def onLockmassCheck(self):
		currentState = self.ui.lockmassCheck.isChecked()
		self.ui.lockmassLab.setVisible(currentState)
		self.ui.lockmassVal.setVisible(currentState)

	def onRawrangCheck(self):
		currentState = self.ui.rawrangeCheck.isChecked()
		self.ui.rawrangeStLab.setVisible(currentState)
		self.ui.rawrangeStVal.setVisible(currentState)
		self.ui.rawrangeEnLab.setVisible(currentState)
		self.ui.rawrangeEnVal.setVisible(currentState)

	def onRawProcess(self):
		params = {}
		params["smoothing"] = None
		if self.ui.rawsmoothCheck.isChecked():
			params["smoothing"] = float(self.ui.rawsmoothVal.text)

		params["lockmass"] = None
		if self.ui.lockmassCheck.isChecked():
			params["lockmass"] = float(self.ui.lockmassVal.text)

		params["range"] = [self.logic.raw_range[0], self.logic.raw_range[1]]
		if self.ui.rawrangeCheck.isChecked():
			params["range"] = [float(self.ui.rawrangeStVal.text), float(self.ui.rawrangeEnVal.text)]

		params["n_ions"] = int(self.ui.rawNIonVal.text)
		params["decimal_ions"] = int(self.ui.rawmzResVal.text)

		process_done = self.logic.raw_processing(params)
		if process_done:
			self.visRenormalize()
			self.logic.heatmap_display()
			self.populateMzLists()
			info = self.logic.getDataInformation()
			self.ui.rawProcessInfo.setText(info)


	def onTextFileSelect(self):
		file_path = self.textFileSelect()
		if file_path:
			self.ui.ImportlineEdit.setText(file_path)
			self.ui.ImportlineEdit.setToolTip(file_path)
			self.onTextFileLoad()

	def textFileSelect(self):
		fileExplorer = qt.QFileDialog()
		
		if self.AppMode==0:
			filePaths = fileExplorer.getOpenFileName(None, "Import MSI data", "", "Structured CSV (*.csv);;Hierarchical HDF5 (*.h5);;Waters DESI Text (*.txt);;Continuous imzML (*.imzml);;All Files (*)")
		elif self.AppMode==1:
			filePaths = fileExplorer.getOpenFileName(None, "Import Image Embeddings", "", "NumPy array (*.npy);;Hierarchical HDF5 (*.h5);;Structured CSV (*.csv);;All Files (*)")

		return filePaths
	
	def onTextFileLoad(self):
		file_load = self.logic.textFileLoad(self.ui.ImportlineEdit.text)
		if file_load:
			tic_normalized = self.visRenormalize()
			info = self.logic.getDataInformation() if tic_normalized else 'Error in TIC Normalization'
		else:
			info = 'Error in File Load. Please check the data and try again.'

		self.dataInfo = info
		self.ui.dataInformation.setText(info)
		self.logic.heatmap_display()
		# if self.AppMode==0:
		# 	self.logic.heatmap_display()
		# elif self.AppMode==1:
		# 	self.logic.pca_display()
		self.populateMzLists()
  
	def onHistoSelect(self):
		histoPath = self.HistofileSelect()
		if histoPath:
			self.ui.HistoLineEdit.setText(histoPath)
			self.ui.HistoLineEdit.setToolTip(histoPath)
			self.onloadHisto()
			
	def HistofileSelect(self):
		fileExplorer = qt.QFileDialog()

		if self.AppMode==0:
			filePath = fileExplorer.getOpenFileName(None, "Open pathology image", "", "Image Files (*.png *.tif* *.jpg *.jpeg);;All Files (*)")
		elif self.AppMode==1:
			filePath = fileExplorer.getOpenFileName(None, "Open image", "", "Image Files (*.png *.tif* *.jpg *.jpeg);;All Files (*)")
		
		return filePath
	
	def onloadHisto(self):
		# when load histopathology is selected runs load histopathology
		self.logic.loadHistopathology(self.ui.HistoLineEdit.text)


	def onREIMSSelect(self):
		file_info = self.logic.REIMSSelect()
		if file_info:
			self.ui.filenameREIMSBrowser.setText(f'{file_info[0]}{file_info[1]}')

	def onREIMSLoad(self):
		# on the text file load runs the text file load and shows the confirmation button

		file_load = self.logic.REIMSLoad(self.ui.filenameREIMSBrowser.toPlainText())
		if file_load:
			tic_normalized = self.visRenormalize()
			info = self.logic.getREIMSInfo() if tic_normalized else 'Error in TIC Normalization'
		else:
			info = 'Error in File Load. Please check the data and try again.'

		self.dataInfo = info
		self.ui.REIMSInformation.setText(info)
		self.logic.heatmap_display()
		self.populateMzLists()

	def onExport(self):
		print('Export MSI data...')
		fileExplorer = qt.QFileDialog()		
		defaultSave = self.logic.savenameBase
		savepath = fileExplorer.getSaveFileName(None, "Export", defaultSave, "Structured CSV (*.csv);;Hierarchical HDF5 (*.h5)")
		print(savepath)
		self.logic.MSIExport(savepath)

	### Visualization tab
	def visRenormalize(self):
		spec_norm_method = self.ui.visNorm_spectra.currentText
		ion_norm_method = self.ui.visNorm_ions.currentText
		norm_flag = self.logic.normalize(spec_norm_method, ion_norm_method)
		return norm_flag
	
	def onSpectrumPlot(self):
		self.logic.spectrum_plot()
		return True
	
	def onSimHeatmap(self):
		volumeNode = self.logic.point_similarity_heatmap()
		sliderWidget = self.ui.simHeatmapSlider
		interactor = SimHeatmapThresholdOverlay(sliderWidget, volumeNode)


	def onPartialPCAButton(self):
		# get all ROIs
		all_rois = slicer.mrmlScene.GetNodesByClass("vtkMRMLMarkupsROINode")
		# select the last ROI created
		roi = all_rois.GetItemAsObject(all_rois.GetNumberOfItems()-1)
		# get roi bounds
		# bounds = np.zeros((6,1))
		bounds = [0.0] * 6
		roi.GetBounds(bounds)
		# error checking to ensure it is within image bounds
		max_x = min(self.logic.dim_x, np.ceil(abs(bounds[0])))
		min_x = 0 if bounds[1] > 0 else np.ceil(-bounds[1])
		max_y = min(self.logic.dim_y, np.ceil(abs(bounds[2])))
		min_y = 0 if bounds[3] > 0 else np.ceil(-bounds[3])

		# print(f'partial PCA region: ({int(min_y)}, {int(max_y)}) ({int(min_x)}, {int(max_x)})')
		
		# logic processes pca in the ROI
		self.logic.partial_pca_display((int(min_y), int(min_x)), (int(max_y), int(max_x)), 
								 extend=self.ui.pcaExtendCheckbox.isChecked())
		
		info = self.logic.LoadingsRank()
		info = f'Local Contrast \nregion ({int(min_y)},{int(max_y)}), ({int(min_x)},{int(max_x)})\n\n' + info
		self.ui.LoadingsInfo.setText(info)
		print(info)

	def onROIforLocalContrast(self):
		# Remove any existing ROI for local contrast
		for node in slicer.util.getNodesByClass("vtkMRMLMarkupsROINode"):
			if node.GetName() == " ":
				slicer.mrmlScene.RemoveNode(node)
				break  # assuming only one exists

		# Create a new Markups ROI node
		roiNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsROINode", " ")

		# Configure display properties (transparent faces, visible edges)
		roiDisplayNode = roiNode.GetDisplayNode()
		roiDisplayNode.SetHandlesInteractive(True)
		roiDisplayNode.SetFillOpacity(0.0)      # Transparent box
		roiDisplayNode.SetOutlineOpacity(1.0)   # Visible edges

		# Set it as the active list for placement
		slicer.modules.markups.logic().SetActiveListID(roiNode)

		# Enable placement mode
		interactionNode = slicer.app.applicationLogic().GetInteractionNode()
		interactionNode.SetCurrentInteractionMode(interactionNode.Place)
		interactionNode.SwitchToSinglePlaceMode()

	def onROIContrast(self):
		# get all ROIs
		segmentationNode = slicer.mrmlScene.GetFirstNodeByClass('vtkMRMLSegmentationNode')
		segments = segmentationNode.GetSegmentation()
		n_segments = segments.GetNumberOfSegments()
		label_mask = np.zeros((self.logic.dim_y, self.logic.dim_x), dtype='bool')

		for i in range(n_segments):
			segment_id = segments.GetNthSegmentID(i)
			try:
				a = slicer.util.arrayFromSegmentBinaryLabelmap(segmentationNode, segment_id)
				a = a*(i+1)
			except:
				referenceVolume = [x for x in slicer.mrmlScene.GetNodesByClass('vtkMRMLScalarVolumeNode') if 'tic' in x.GetName()][0]
				a = slicer.util.arrayFromSegmentBinaryLabelmap(segmentationNode, segment_id, referenceVolume)
				a = a*(i+1)
			label_mask = label_mask + a
		
		# logic processes pca in the ROI
		self.logic.roi_pca_display(label_mask, extend=self.ui.roiCintrastExtend.isChecked())
		
		info = self.logic.LoadingsRank()
		info = 'ROI Contrast \n\n' + info
		self.ui.LoadingsInfo.setText(info)


	def onROIContrastLDA(self):
		# get all ROIs
		segmentationNode = slicer.mrmlScene.GetFirstNodeByClass('vtkMRMLSegmentationNode')
		segments = segmentationNode.GetSegmentation()
		n_segments = segments.GetNumberOfSegments()
		label_mask = np.zeros((self.logic.dim_y, self.logic.dim_x), dtype='bool')

		for i in range(n_segments):
			segment_id = segments.GetNthSegmentID(i)
			try:
				a = slicer.util.arrayFromSegmentBinaryLabelmap(segmentationNode, segment_id)
				a = a*(i+1)
			except:
				referenceVolume = [x for x in slicer.mrmlScene.GetNodesByClass('vtkMRMLScalarVolumeNode') if 'tic' in x.GetName()][0]
				a = slicer.util.arrayFromSegmentBinaryLabelmap(segmentationNode, segment_id, referenceVolume)
				a = a*(i+1)
			label_mask = label_mask + a
		
		# logic processes pca in the ROI
		self.logic.roi_lda_display(label_mask, extend=self.ui.roiCintrastExtend.isChecked())

	def getSegmentData(self):
		"""get the mask, color, and name of the segmentations"""

		segmentationNode = slicer.mrmlScene.GetFirstNodeByClass('vtkMRMLSegmentationNode')
		segments = segmentationNode.GetSegmentation()
		segment_IDs = segments.GetSegmentIDs()
		segMask, segName, segColor = [], [], []

		for segment_id in segment_IDs:
			segment_mask = slicer.util.arrayFromSegmentBinaryLabelmap(segmentationNode, segment_id) # 1,dim_uy,dim_x
			segment_color = segments.GetSegment(segment_id).GetColor() # r,g,b
			segment_name = segments.GetSegment(segment_id).GetName()

			segMask.append(segment_mask)
			segName.append(segment_name)
			segColor.append(segment_color)

		return segMask, segName, segColor, segmentationNode, segment_IDs


	def onROISimilarity(self):
		segMask, segName, segColor, _, _ = self.getSegmentData()
		similarity_threshold = self.ui.similarityThreshold.value
		_ = self.logic.roi_similarity_map(segMask, segName, segColor, similarity_threshold)

	def onROIExpand(self):
		segMask, segName, segColor, segmentationNode, segID = self.getSegmentData()
		similarity_threshold = self.ui.similarityThreshold.value
		roi_expansion_ind = self.logic.roi_similarity_map(segMask, segName, segColor, similarity_threshold)
		for ind in range(len(segMask)):
			mask = segMask[ind]
			mask[roi_expansion_ind==ind] = 1
			slicer.util.updateSegmentBinaryLabelmapFromArray(
				mask,
				segmentationNode,
				segID[ind])

	def onSegmentVisibility(self):
		lm = slicer.mrmlScene.GetFirstNodeByClass('vtkMRMLSegmentationNode')
		vis = lm.GetDisplayVisibility()
		if vis:
			lm.SetDisplayVisibility(0)
		else:
			lm.SetDisplayVisibility(1)

	def onPCAButton(self):
		# displays the pca image
		self.logic.pca_display()
		
		info = self.logic.LoadingsRank()
		info = 'Global Contrast \n\n' + info
		self.ui.LoadingsInfo.setText(info)

	def onSaveProjection(self):
		fileExplorer = qt.QFileDialog()
		defaultSave = self.logic.savenameBase+"_PCA.pkl"
		savepath = fileExplorer.getSaveFileName(None, "Save PCA projection", defaultSave, "Pickle Files (*.pkl);;All Files (*)")
		
		pixel_norm_method = self.ui.visNorm_spectra.currentText
		feature_norm_method = self.ui.visNorm_ions.currentText
		self.logic.pca_export(savepath, pixel_norm_method, feature_norm_method)
		
		print(savepath)

	def onLoadProjection(self):
		fileExplorer = qt.QFileDialog()
		filePath = fileExplorer.getOpenFileName(None, "Load PCA projection", "", "Pickle Files (*.pkl);;All Files (*)")
		print(filePath)
		self.logic.pca_import(filePath)
		pass

	def onNLVisMethod(self, text):
		if text=="UMAP":
			self.ui.NLVisLabel1.setText("n_neighbors")
			self.ui.NLVisLabel1.setToolTip("2 - 200")
			self.ui.NLVisParam1.setText("15")
			self.ui.NLVisLabel2.setText("min_dist")
			self.ui.NLVisLabel2.setToolTip("0.0 - 0.99")
			self.ui.NLVisParam2.setText("0.1")
			
		elif text=="t-SNE":
			self.ui.NLVisLabel1.setText("perplexity")
			self.ui.NLVisLabel1.setToolTip("5 - 50")
			self.ui.NLVisParam1.setText("30")
			self.ui.NLVisLabel2.setText("early_exaggeration")
			self.ui.NLVisLabel2.setToolTip("4 - 20")
			self.ui.NLVisParam2.setText("12")
	
	def onUMAPVis(self):
		method = self.ui.NLVisMethod.currentText
		param1 = float(self.ui.NLVisParam1.text)
		param2 = float(self.ui.NLVisParam2.text)
		self.logic.nonlinear_display(method, param1, param2)
  
	def populateMzLists(self):
		# adds the m/z values to m/z lists for the color channels
		self.ui.mzlist.clear()
		self.ui.mzlist_2.clear()
		self.ui.mzlist_3.clear()
		self.ui.mzlist_4.clear()
		self.ui.mzlist_5.clear()
		self.ui.mzlist_6.clear()
		self.ui.mzlist_7.clear()
		self.ui.singleIonMzList.clear()

		self.ui.mzlist.addItem("None")
		self.ui.mzlist_2.addItem("None")
		self.ui.mzlist_3.addItem("None")
		self.ui.mzlist_4.addItem("None")
		self.ui.mzlist_5.addItem("None")
		self.ui.mzlist_6.addItem("None")
		self.ui.mzlist_7.addItem("None")
		self.ui.singleIonMzList.addItem("None")
		
		all_mz = list(self.logic.mz)
		for mz in all_mz:
			self.ui.mzlist.addItem(mz)
			self.ui.mzlist_2.addItem(mz)
			self.ui.mzlist_3.addItem(mz)
			self.ui.mzlist_4.addItem(mz)
			self.ui.mzlist_5.addItem(mz)
			self.ui.mzlist_6.addItem(mz)
			self.ui.mzlist_7.addItem(mz)
			self.ui.singleIonMzList.addItem(mz)

		
	def selectedMultiIon(self):
		# runs the valudation for all of the color channels
		self.logic.multiIonVisualization([["white",self.ui.mzlist.currentText], 
									["red",self.ui.mzlist_2.currentText],
									["green",self.ui.mzlist_3.currentText],
									["blue",self.ui.mzlist_4.currentText],
									["yellow",self.ui.mzlist_7.currentText], 
									["magenta",self.ui.mzlist_5.currentText],
									["cyan",self.ui.mzlist_6.currentText]])
		
	def selectedSingleIon(self):
		# runs the valudation for all of the color channels
		self.logic.singleIonVisualization((self.ui.singleIonMzList.currentText), 
									self.ui.singleIonHeatmapList.currentText)

	def onAbundanceThumbnail(self):
		self.logic.ViewAbundanceThumbnail()

	def onContrastThumbnail(self):
		self.logic.ViewContrastThumbnail()

	def onClusterButton(self):
		n_clusters = int(self.ui.nCluster.currentText)
		cluster_colors = self.logic.VisCluster(n_clusters)
		# self.ClearClusterTable()
		self.ui.ClusterInd.clear()
		for i in range(n_clusters):
			item = qt.QStandardItem(f'cluster {i+1}')
			color = tuple(int(component * 255) for component in cluster_colors[i])
			item.setForeground(qt.QColor(*color))
			self.ui.ClusterInd.model().appendRow(item)

	def ClearClusterTable(self):
		self.ui.ClusterTable.setRowCount(1)
		for j in range(5):
			self.ui.ClusterTable.setItem(0, j, qt.QTableWidgetItem(' '))
		self.ui.ClusterTable.resizeColumnsToContents()  # Adjust column widths
		self.ui.ClusterTable.resizeRowsToContents()     # Adjust row heights
		self.ui.ClusterTable.sortItems(1, qt.Qt.DescendingOrder)

	def onClusterThumbnail(self):
		self.TableAndThumbnail("cluster")

	def onSimThumbnail(self):
		self.TableAndThumbnail("similarity")

	def TableAndThumbnail(self, mode):
		if mode=="cluster":
			clusterText = self.ui.ClusterInd.currentText
			cluster_ind = int(clusterText.split(' ')[-1])-1
			cluster_invert = self.ui.invertCluster_checkBox.isChecked()
			volcano_mz, dice_score, volcano_fc, volcano_pval, pearson_corr = \
				self.logic.ViewTableThumbnail([cluster_ind, cluster_invert], mode)
		elif mode=="similarity":
			sim_thresh_value = self.ui.simHeatmapSlider.value
			volcano_mz, dice_score, volcano_fc, volcano_pval, pearson_corr = \
				self.logic.ViewTableThumbnail(sim_thresh_value, mode)

		# Build DataFrame
		clusterIons = pd.DataFrame({
			'm/z': volcano_mz,
			'Pearson correlation': np.round(pearson_corr, 4),
			'FC [log]': np.round(volcano_fc, 4),
			'p-value [-log]': np.round(volcano_pval, 4),
			'Dice score': np.round(dice_score, 4),
		})
		if self.AppMode==1:
			clusterIons = clusterIons.rename(columns={"m/z": "feature"})
		# clusterIons.to_csv(self.logic.savenameBase+"_test.csv")

		lm = slicer.app.layoutManager()
		layoutNode = lm.layoutLogic().GetLayoutNode()

		# Ensure the layout
		customLayoutId = 80
		customLayout = """
		<layout type="vertical" split="true">
		<item>
			<view class="vtkMRMLSliceNode" singletontag="Yellow"/>
		</item>
		<item>
			<view class="vtkMRMLTableViewNode" singletontag="ThumbnailTable"/>
		</item>
		</layout>
		"""
		# Add/refresh description
		if layoutNode.GetLayoutDescription(customLayoutId) != customLayout:
			layoutNode.AddLayoutDescription(customLayoutId, customLayout)
		lm.setLayout(customLayoutId)

		# TableView node (singleton tag "ThumbnailTable")
		tableViewNode = slicer.mrmlScene.GetSingletonNode("ThumbnailTable", "vtkMRMLTableViewNode")
		# Fallback if needed:
		if not tableViewNode:
			tvs = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
			tableViewNode = tvs[0] if tvs else None

		# Create/clear the table node and fill it
		tableNode = slicer.mrmlScene.GetFirstNodeByName("ClusterIons")
		if not tableNode:
			tableNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLTableNode', 'ClusterIons')
		else:
			tableNode.RemoveAllColumns()

		# for col in clusterIons.columns:
		# 	arr = vtk.vtkVariantArray()
		# 	arr.SetName(str(col))
		# 	for v in clusterIons[col].values:
		# 		# keep numeric where possible; vtkVariant will wrap it
		# 		arr.InsertNextValue(vtk.vtkVariant(v))
		# 	tableNode.AddColumn(arr)

		for col in clusterIons.columns:
			# convert to string to handle mixed data type
			arr = vtk.vtkStringArray()
			arr.SetName(str(col))
			for v in clusterIons[col].values:
				arr.InsertNextValue(str(v))
			tableNode.AddColumn(arr)

		tableNode.SetUseColumnTitleAsColumnHeader(True)
		tableNode.SetLocked(True)
		tableNode.Modified()  # nudge MRML/UI

		# Bind table to the Table view
		if tableViewNode:
			tableViewNode.SetTableNodeID(tableNode.GetID())

		# Process pending UI events
		slicer.app.processEvents()


	### Blender
	def _setBlendWidgetsEnabled(self, enabled):
		for w in self.ui.blendGroupBox.findChildren(qt.QWidget):
			w.setEnabled(enabled)

	def _onBlendEnabledChanged(self, enabled):
		self._setBlendWidgetsEnabled(enabled)

		if not enabled:
			self._disableBlend()
			if hasattr(self, "_blendSweepTimer"):
				self._blendSweepTimer.stop()
			return

		# enable
		self.ui.overlayRadioButton.checked = True
		self.ui.manualBlendSliderRadioButton.checked = True
		
		if self._blendCtrl is None:
			# Create logic controller for Red view
			self._blendCtrl = self.logic.SliceBlendController(sliceViewName="Red",
															outputName="Blended",
															wipeDirection="horizontal")
		self._applyBlendInputs()
		self._onBlendSliderChanged(self.ui.blendSlider.value)

	def _disableBlend(self):
		# Just drop controller reference; it does not hold Qt connections.
		# Also restore overlay-like display (so user isn't stuck seeing the output volume)
		if self._blendCtrl is not None:
			bg = self.ui.bgVolumeSelector.currentNode()
			fg = self.ui.fgVolumeSelector.currentNode()
			self._blendCtrl.setVolumes(bg, fg)
			self._blendCtrl.setMode("overlay")
			self.ui.overlayRadioButton.checked = True
			self._blendCtrl = None

	def _scheduleBlendApply(self):
		if not self.ui.blendGroupBox.isChecked():
			return
		# debounce 80ms so rapid UI changes don't thrash
		self._blendDebounce.start(80)

	def _applyBlendInputs(self):
		if self._blendCtrl is None:
			return

		bg = self.ui.bgVolumeSelector.currentNode()
		fg = self.ui.fgVolumeSelector.currentNode()
		self._blendCtrl.setVolumes(bg, fg)

		# Determine mode + wipe direction from radio buttons
		if self.ui.overlayRadioButton.checked:
			mode = "overlay"
			direction = "horizontal"  # unused in overlay
		elif self.ui.verticalWipeRadioButton.checked:
			mode = "wipe"
			direction = "vertical"
		else:
			# existing wipeRadioButton = horizontal wipe
			mode = "wipe"
			direction = "horizontal"

		# Apply
		self._blendCtrl.wipeDirection = direction
		self._blendCtrl.setMode(mode)

		# Apply current slider immediately (opacity in overlay, wipe fraction in wipe)
		self._blendCtrl.updateFromSlider(self.ui.blendSlider, self.ui.blendSlider.value)


	def _onBlendModeToggled(self, checked):
		if not checked:
			return
		if not self.ui.blendGroupBox.isChecked():
			return

		# Pause sweep during mode switching/capture to avoid race
		wasSweeping = hasattr(self, "_blendSweepTimer") and self._blendSweepTimer.isActive()
		if wasSweeping:
			self._blendSweepTimer.stop()

		self._scheduleBlendApply()

		# Resume sweep shortly after the mode switch has settled
		if wasSweeping and self.ui.sweepBlendSliderRadioButton.checked:
			qt.QTimer.singleShot(120, lambda: (
				self.ui.blendGroupBox.isChecked()
				and self.ui.sweepBlendSliderRadioButton.checked
				and self._blendSweepTimer.start()
			))


	def _onBlendSliderChanged(self, v):
		if not self.ui.blendGroupBox.isChecked() or self._blendCtrl is None:
			return
		self._blendCtrl.updateFromSlider(self.ui.blendSlider, v)

	def _setupBlendSliderSweepUI(self):
		# --- Separate radio button group for sweep/manual ---
		self._blendSweepButtonGroup = qt.QButtonGroup(self.ui.blendGroupBox)
		self._blendSweepButtonGroup.setExclusive(True)
		self._blendSweepButtonGroup.addButton(self.ui.manualBlendSliderRadioButton)
		self._blendSweepButtonGroup.addButton(self.ui.sweepBlendSliderRadioButton)

		# Default
		self.ui.manualBlendSliderRadioButton.checked = True

		# --- Sweep engine ---
		self._blendSweepTimer = qt.QTimer()
		self._blendSweepTimer.setInterval(50)  # ms tick
		self._blendSweepTimer.timeout.connect(self._onBlendSweepTick)

		self._blendSweepDir = +1
		self._blendSweepSecondsOneWay = 1.5  # min->max (or max->min) takes 2 seconds

		# React when a radio becomes checked (avoid double-trigger)
		self.ui.manualBlendSliderRadioButton.toggled.connect(self._onBlendSweepModeToggled)
		self.ui.sweepBlendSliderRadioButton.toggled.connect(self._onBlendSweepModeToggled)

	def _onBlendSweepModeToggled(self, checked):
		# Only act on the button that became checked
		if not checked:
			return

		# If blending is disabled, never run sweep
		if not self.ui.blendGroupBox.isChecked():
			self._blendSweepTimer.stop()
			return

		if self.ui.sweepBlendSliderRadioButton.checked:
			self._blendSweepDir = +1
			self._blendSweepTimer.start()
		else:
			self._blendSweepTimer.stop()

	def _onBlendSweepTick(self):
		# Stop conditions
		if (not self.ui.blendGroupBox.isChecked()) or (not self.ui.sweepBlendSliderRadioButton.checked):
			self._blendSweepTimer.stop()
			return

		s = self.ui.blendSlider

		mn = float(s.minimum)
		mx = float(s.maximum)
		if mx <= mn:
			return

		# We want: one-way sweep takes self._blendSweepSecondsOneWay seconds
		dt = float(self._blendSweepTimer.interval) / 1000.0
		step = (mx - mn) * (dt / float(self._blendSweepSecondsOneWay))

		v = float(s.value) + self._blendSweepDir * step

		# Bounce at ends (clamp + reverse direction)
		if v >= mx:
			v = mx
			self._blendSweepDir = -1
		elif v <= mn:
			v = mn
			self._blendSweepDir = +1

		# Setting slider drives your existing valueChanged -> blend update
		s.value = v

	### Blender


	### Dataset generation

	def landmark(self):
		# swtiches module to landmark registration
		pluginHandlerSingleton = slicer.qSlicerSubjectHierarchyPluginHandler.instance()
		pluginHandlerSingleton.pluginByName('Default').switchToModule("LandmarkRegistration")

	def showSegmentEditor(self):
		segSelect1Node = self.ui.segVolCombo1.currentNode()
		segSelect2Node = self.ui.segVolCombo2.currentNode()

		if segSelect1Node:
		
			slicer.util.selectModule("SegmentEditor")

			# set master volume and geometry
			sourceVolumeNode = segSelect1Node
			segmentEditorNode = slicer.util.getNodesByClass('vtkMRMLSegmentEditorNode')[0]
			segmentationNode = slicer.util.getNodesByClass('vtkMRMLSegmentationNode')[0]
			segmentationNode.SetReferenceImageGeometryParameterFromVolumeNode(sourceVolumeNode)
			segmentEditorWidget = slicer.qMRMLSegmentEditorWidget()
			segmentEditorWidget.setMRMLScene(slicer.mrmlScene)
			segmentEditorWidget.setMRMLSegmentEditorNode(segmentEditorNode)
			segmentEditorWidget.setSegmentationNode(segmentationNode)
			segmentEditorWidget.setSourceVolumeNode(sourceVolumeNode)

			# set the display view and link views
			RedCompNode = slicer.util.getNode("vtkMRMLSliceCompositeNodeRed")
			RedCompNode.SetBackgroundVolumeID(segSelect1Node.GetID())
			RedNode = slicer.util.getNode("vtkMRMLSliceNodeRed")
			RedNode.SetOrientation("Axial")

			if not segSelect2Node:
				slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUpRedSliceView)

			else:
				YellowCompNode = slicer.util.getNode("vtkMRMLSliceCompositeNodeYellow")
				YellowCompNode.SetBackgroundVolumeID(segSelect2Node.GetID())
				YellowNode = slicer.util.getNode("vtkMRMLSliceNodeYellow")
				YellowNode.SetOrientation("Axial")
				RedCompNode.SetLinkedControl(True)
				YellowCompNode.SetLinkedControl(True)
				slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutSideBySideView)
				
			slicer.util.resetSliceViews()

  
	def onCSVconnect(self, meta_only=False):
		fileExplorer = qt.QFileDialog()
		# defaultSave = self.ui.filenameTextBrowser.toPlainText()[:-4]+'_dataset'
		defaultSave = os.path.splitext(self.ui.ImportlineEdit.text)[0]
		if meta_only:
			defaultSave += '_metadata'
		else:
			defaultSave += '_dataset'
		
		savepath = fileExplorer.getSaveFileName(None, "Save aligned dataset", defaultSave, "CSV Files (*.csv);;All Files (*)")
		
		retstr = self.logic.csvGeneration(savepath, meta_only)
		self.ui.csvcreateTextBrowser.setText(retstr)

		if meta_only:
			self.logic.segmentationSave(savepath)
	
	def onSaveScene(self):
		print('saving the project...')
		fileExplorer = qt.QFileDialog()
		# defaultSave = self.ui.filenameTextBrowser.toPlainText()[:-4]+'_project'
		defaultSave = self.ui.ImportlineEdit.text[:-4]+'_project'
		savepath = fileExplorer.getSaveFileName(None, "Save Project", defaultSave, "MSI Project Files (*.mrb);;All Files (*)")
		print(savepath)
		slicer.util.saveScene(savepath)


	def onTabChange(self, index):
		tab_names = ['Home', 'Data', 'Visualization', 'Dataset', 'Alignment', 'Preprocessing', 'Statistical', 'AI training', 'AI report', 'AI deployment', 'Identification']
		for i in range(self.ui.tabWidget.count):
			self.ui.tabWidget.setTabText(i, "")     
		self.ui.tabWidget.setTabText(index, tab_names[index])
		print('selected tab:',index, self.ui.tabWidget.tabText(index))
		if index==9:
			# self.updateDepVisList()
			self.updateDepSegList()
		# elif index==3:
		# 	self.updateVolumeList()
	
	def onModuleChange(self):
		# self.updateDepVisList()
		self.updateDepSegList()
		# self.updateVolumeList()

	# def updateVolumeList(self):
	# 	self.ui.segVollist1.clear()
	# 	self.ui.segVollist1.addItem('None')
	# 	self.ui.segVollist2.clear()
	# 	self.ui.segVollist2.addItem('None')
	# 	volumeNodes = slicer.util.getNodesByClass('vtkMRMLScalarVolumeNode')
	# 	for volumeNode in volumeNodes:
	# 		self.ui.segVollist1.addItem(volumeNode.GetName())
	# 		self.ui.segVollist2.addItem(volumeNode.GetName())

	def updateDepSegList(self):
		segmentationNode = slicer.util.getNodesByClass('vtkMRMLSegmentationNode')
		if len(segmentationNode)>0:
			segmentationNode = segmentationNode[0]
			segmentation = segmentationNode.GetSegmentation()
			segIDs = segmentation.GetSegmentIDs()
			segNames = [segmentation.GetSegment(segID).GetName() for segID in segIDs]

			self.ui.depSegListCombo.clear()
			for segName in segNames:
				self.ui.depSegListCombo.addItem(segName)
	
	#----- Robert Peak Labeling Running/Functions to do so -------
	def onRadioButtonClicked(self):
		"""Returns True if 'find all' is selected, False if 'find closest' is selected."""
		if self.ui.findallcandidatesradioButton.isChecked():
			return True
		else:
			return False
		
	def onUpdateHMDBDatabase(self):
		"""Updates the HMDB database by calling the logic function and displays a message box with the result."""
		db_path = self.logic.default_hmdb_db_path()	
		buttonClicked = True
		result = self.logic.check_and_build_hmdb(db_path, buttonClicked)
		if result:
			slicer.util.infoDisplay("HMDB database updated successfully!")
	
	def onLabelPeaks(self):
		' Function to label the peaks based on the inputted m/z values, tolerance, and adducts. '
		'It also handles the UI updates to show the results in a table and allows for pathway searching. '
		#Gather inputs
		raw_peaks = self.ui.inputtedpeakslineedit.text
		tolerance_str = self.ui.moleculetoleranacelineedit.text
		checked_items = []
		if self.ui.mhposcheckbox_2.isChecked():
			checked_items.append("M+H")
		if self.ui.mnaposcheckbox_2.isChecked():
			checked_items.append("M+Na")
		if self.ui.mkposcheckbox_2.isChecked():
			checked_items.append("M+K")
		if self.ui.mnh4poscheckbox_2.isChecked():
			checked_items.append("M+NH4")
		if self.ui.mhnegcheckbox_2.isChecked():
			checked_items.append("M-H")
		if self.ui.mclnegcheckbox_2.isChecked():
			checked_items.append("M+Cl")
		if self.ui.mfnegcheckbox_2.isChecked():
			checked_items.append("M-F")
		if self.ui.mch3coonegcheckbox_2.isChecked():
			checked_items.append("M+CH3COO")
		if self.ui.noadductcheckbox_2.isChecked():
			checked_items.append("Neutral")
		adducts_text = ",".join(checked_items)
		#use_broad_pathways = self.ui.nonspecificpathwayscheckbox.isChecked()
		
		search_all = self.onRadioButtonClicked()
		tol_unit = self.ui.toleranceunitcombobox.currentText
		database_unit = self.ui.databasecombobox.currentText

		# Validation and error handling
		if not raw_peaks:
			slicer.util.errorDisplay("Please enter at least one peak m/z value.")
			return
		try:
			tolerance = float(tolerance_str)
		except ValueError:
			slicer.util.errorDisplay("Tolerance must be a valid number!")
			return
		if not adducts_text:
			slicer.util.errorDisplay("Please select at least one adduct.")
			return
		
		# Update UI to show it's working (API calls take time)
		self.ui.displaypatwaystextbrowser.show()
		slicer.app.processEvents()
		
		progress_dialog = slicer.util.createProgressDialog(labelText="Finding Molecules...", maximum=100)
		def update_progress(message, percentage):
			progress_dialog.labelText = message
			progress_dialog.setValue(percentage)
			slicer.app.processEvents()

		try:
			# Run the molecule matching logic
			results_df = self.logic.run_molecule_matching(
				raw_peaks, tolerance, adducts_text, tol_unit, database_unit, search_all, update_progress
			)
			self.current_results_df = results_df

			if results_df is None or (isinstance(results_df, str)) or results_df.empty:
				self.ui.moleculesTableWidget.setRowCount(0)
				slicer.util.infoDisplay("No molecule matches found for the given peaks and parameters.")
				return
			results_df = results_df.sort_values(by=['Searched_m/z', 'DELTA'])
			results_df = results_df.reset_index(drop=True)
			self.current_results_df = results_df
			self.current_pathway_df = None

			self.ui.displaypatwaystextbrowser.setText("") # Clear out old results
			
			self.ui.moleculesTableWidget.show()
			self.ui.searchPathwaysButton.show()
			self.ui.moleculesTableWidget.setHorizontalHeaderLabels(['Select', 'Searched m/z', 'Adduct', 'Molecule', 'Source ID', 'KEGG ID', f'Error ({tol_unit})'])
			self.ui.moleculesTableWidget.setRowCount(0)
			slicer.app.processEvents() 

			# Populate the Table
			self.ui.moleculesTableWidget.setRowCount(len(results_df))
			for i, row in results_df.iterrows():
				chk_item = qt.QTableWidgetItem()
				chk_item.setFlags(qt.Qt.ItemIsUserCheckable | qt.Qt.ItemIsEnabled)
				chk_item.setCheckState(qt.Qt.Checked) 
				chk_item.setData(qt.Qt.UserRole, i)

				self.ui.moleculesTableWidget.setItem(i, 0, chk_item)
				self.ui.moleculesTableWidget.setItem(i, 1, qt.QTableWidgetItem(str(row.get('Searched_m/z', ''))))
				self.ui.moleculesTableWidget.setItem(i, 2, qt.QTableWidgetItem(str(row.get('Adduct', ''))))
				self.ui.moleculesTableWidget.setItem(i, 3, qt.QTableWidgetItem(str(row.get('COMMON_NAME', ''))))
				self.ui.moleculesTableWidget.setItem(i, 4, qt.QTableWidgetItem(str(row.get('Source ID', ''))))

				kegg_id = str(row.get('KEGG_ID', ''))
				# If KEGG ID is valid, make it a clickable link to the KEGG entry; otherwise just display the text
				if kegg_id and kegg_id.startswith('C'):
					link_label = qt.QLabel(f'<a href="https://www.kegg.jp/entry/{kegg_id}" style="color: #2980b9; text-decoration: underline;">{kegg_id}</a>')
					link_label.setTextFormat(qt.Qt.RichText)
					link_label.setTextInteractionFlags(qt.Qt.TextBrowserInteraction)
					link_label.setAlignment(qt.Qt.AlignVCenter | qt.Qt.AlignLeft)
					link_label.setStyleSheet("margin-left: 5px;")
					link_label.linkActivated.connect(self.onTableLinkClicked)
					
					self.ui.moleculesTableWidget.setCellWidget(i, 5, link_label)
				else:
					self.ui.moleculesTableWidget.setItem(i, 5, qt.QTableWidgetItem(kegg_id))

				self.ui.moleculesTableWidget.setItem(i, 6, qt.QTableWidgetItem(str(row.get('DELTA', ''))))
				
		finally:
			progress_dialog.close()

	def onSearchPathways(self):
		"""Triggered when the user clicks 'Search Pathways' after selecting molecules.
		It gathers the selected molecules, runs the pathway search logic, and updates the UI with results."""
		self.ui.displaypatwaystextbrowser.clear()
		selected_indices = []
		
		# Gather all rows where the checkbox is ticked
		for row in range(self.ui.moleculesTableWidget.rowCount):
			chk_item = self.ui.moleculesTableWidget.item(row, 0)
			if chk_item is not None and chk_item.checkState() == qt.Qt.Checked:
				df_idx = chk_item.data(qt.Qt.UserRole)
				selected_indices.append(df_idx)

		if not selected_indices:
			slicer.util.errorDisplay("Please select at least one molecule to search pathways for.")
			return

		filtered_df = self.current_results_df.loc[selected_indices].copy()
		filter_human = self.ui.onlyhumanpathwayscheckbox.isChecked()

		# Update UI to show it's working (API calls take time)
		progress_dialog = slicer.util.createProgressDialog(labelText="Searching Pathways...", maximum=100)
		def update_progress(message, percentage):
			progress_dialog.labelText = message
			progress_dialog.setValue(percentage)
			slicer.app.processEvents()

		# Get Final Pathway DataFrame and update UI
		try:
			final_df = self.logic.run_pathway_search(filtered_df, filter_human, update_progress)
			self.current_pathway_df = final_df
			self.ui.displaypatwaystextbrowser.show()
			
			slicer.app.processEvents()

			if final_df is None or final_df.empty or 'Pathway_Name' not in final_df.columns:
				self.ui.displaypatwaystextbrowser.setText("No pathways found for the selected molecules.")
			else:
				self.renderHtmlPathwayResults(final_df, self.ui.moleculetoleranacelineedit.text, self.ui.toleranceunitcombobox.currentText)

		finally:
			progress_dialog.close()

	def renderHtmlPathwayResults(self, results_df, tolerance, tol_unit):
		"""Generates and sets the HTML layout for the final pathway results."""
		output_html = f"""
		<html>
		<head>
		<style>
			body {{ font-family: Arial, sans-serif; color: #333; }}
			h3 {{ color: #2c3e50; margin-bottom: 2px; }}
			h4 {{ color: #1a5276; margin-top: 20px; margin-bottom: 5px; border-bottom: 2px solid #1a5276; padding-bottom: 3px; }}
			p {{ margin-top: 0px; color: #555; }}
			table {{ border-collapse: collapse; width: 100%; margin-bottom: 15px; }}
			th, td {{ padding: 6px 10px; text-align: left; border-bottom: 1px solid #d4d4d4; }}
			th {{ background-color: #e0e0e0; font-weight: bold; color: #333; }}
			tr:nth-child(even) {{ background-color: #f9f9f9; }}
		</style>
		</head>
		<body>
			<h3>Matching & Pathways Complete</h3>
			<p><b>Searched Tolerance:</b> {tolerance} {tol_unit}</p>
		"""
		
		display_cols = ['Searched_m/z', 'Adduct', 'COMMON_NAME', 'Source ID', 'KEGG_ID', 'DELTA', 'Pathway_Name', 'Pathway_ID']
		
		for mz_val, group_df in results_df.groupby('Searched_m/z'):
			display_df = group_df[[col for col in display_cols if col in group_df.columns]].copy()
			display_df = display_df.sort_values(by=['Adduct', 'KEGG_ID'])

			# Make KEGG_ID and Pathway_Name clickable links if the relevant information is present
			if 'Pathway_ID' in display_df.columns:
				display_df['Pathway_Name'] = '<a href="https://www.kegg.jp/pathway/' + display_df['Pathway_ID'] + '" style="color: #2980b9; text-decoration: none;">' + display_df['Pathway_Name'] + '</a>'
				display_df.drop(columns=['Pathway_ID'], inplace=True)
				
			if 'KEGG_ID' in display_df.columns:
				display_df['KEGG_ID'] = '<a href="https://www.kegg.jp/entry/' + display_df['KEGG_ID'] + '" style="color: #2980b9; text-decoration: none;">' + display_df['KEGG_ID'] + '</a>'

			display_df.rename(columns={
				'Searched_m/z': 'Searched m/z', 'Adduct': 'Adduct', 'COMMON_NAME': 'Molecule', 'Source ID': 'Source ID',
				'KEGG_ID': 'KEGG ID', 'DELTA': f'Error ({tol_unit})', 'Pathway_Name': 'Pathway'
			}, inplace=True)
			
			if 'Searched m/z' in display_df.columns:
				display_df.drop(columns=['Searched m/z'], inplace=True)

			output_html += f"<h4>>> Matches for m/z: {mz_val} <<</h4>\n"
			output_html += display_df.to_html(index=False, border=0, justify='left', escape=False)

		output_html += "</body></html>"
		self.ui.displaypatwaystextbrowser.setText(output_html)
	
	def onExportPeakLabelExcel(self):
		"""Export the pathway results to an excel file, preserving URLs."""
		# Import openpyxl for Excel writing and hyperlink support        
		try:
			import openpyxl
			from openpyxl.styles import Font
		except ModuleNotFoundError:
			slicer.util.pip_install("openpyxl")
			import openpyxl
			from openpyxl.styles import Font
		
        # Check if there is actually data to save
		df_to_export = None
		if hasattr(self, 'current_pathway_df') and self.current_pathway_df is not None and not self.current_pathway_df.empty:
			df_to_export = self.current_pathway_df.copy()

		if df_to_export is None:
			slicer.util.infoDisplay("There are no results to export. Please run a pathway search first.")
			return
		
		# Do not save the LIPID_CLASS column if it exists, as it is not relevant to the user
		df_to_export.drop(columns=['LIPID_CLASS', 'FORMULA'], inplace=True, errors='ignore')

        # Open a "Save As" dialog window
		default_name = "Pathway_Labeling_Results.xlsx"
		file_path = qt.QFileDialog.getSaveFileName(
            None, 
            "Save Results as excel", 
            default_name, 
            "Excel Files (*.xlsx)"
        )

        # If the user clicked "Save" (and didn't hit cancel)
		if file_path:
			try:
                # Force the extension to be .xlsx 
				if not file_path.lower().endswith('.xlsx'):
					file_path = os.path.splitext(file_path)[0] + '.xlsx'

                # Use Pandas ExcelWriter to save the base data
				with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
					df_to_export.to_excel(writer, index=False, sheet_name='Pathways')
                    
                    # Access the underlying workbook to add native links
					workbook = writer.book
					worksheet = workbook['Pathways']
                    
                    # Standard Excel hyperlink styling (Blue and Underlined)
					link_font = Font(color="0563C1", underline="single")
                    
                    # Find which column numbers belong to our IDs (openpyxl is 1-indexed)
					columns = df_to_export.columns.tolist()
					kegg_col_idx = columns.index('KEGG_ID') + 1 if 'KEGG_ID' in columns else None
					path_name_idx = columns.index('Pathway_Name') + 1 if 'Pathway_Name' in columns else None
					path_id_idx = columns.index('Pathway_ID') + 1 if 'Pathway_ID' in columns else None
					hmdb_col_idx = None
					if 'Source ID' in columns:
						hmdb_col_idx = columns.index('Source ID') + 1
					elif 'HMDB_ID' in columns:
						hmdb_col_idx = columns.index('HMDB_ID') + 1

                    # Iterate through the rows to add the hyperlinks directly to the text
					for row in range(2, len(df_to_export) + 2): # Start at 2 to skip the header row
                        
                        # Apply link to KEGG_ID
						if kegg_col_idx:
							kegg_cell = worksheet.cell(row=row, column=kegg_col_idx)
							val = kegg_cell.value
							if pd.notna(val) and str(val).startswith('C'):
								kegg_cell.hyperlink = f"https://www.kegg.jp/entry/{val}"
								kegg_cell.font = link_font
                                
                        # Apply link to Pathway_Name
						if path_name_idx and path_id_idx:
							name_cell = worksheet.cell(row=row, column=path_name_idx)
							id_val = worksheet.cell(row=row, column=path_id_idx).value
							if pd.notna(id_val):
								name_cell.hyperlink = f"https://www.kegg.jp/pathway/{id_val}"
								name_cell.font = link_font
						
						# Apply link to HMDB ID if available
						if hmdb_col_idx:
							hmdb_cell = worksheet.cell(row=row, column=hmdb_col_idx)
							val = hmdb_cell.value
							if pd.notna(val):
								hmdb_cell.hyperlink = f"https://www.hmdb.ca/metabolites/{val}"
								hmdb_cell.font = link_font

                    # Clean up: Delete the Pathway_ID column since the links are now attached to the Name
					if path_id_idx:
						worksheet.delete_cols(path_id_idx)

				slicer.util.infoDisplay(f"Successfully saved Excel file to:\n{file_path}", "Export Complete")
            
			except Exception as e:
				slicer.util.errorDisplay(f"Failed to save Excel file:\n{str(e)}")

	def onLoadMzValuesCsv(self):
		"""Opens a file dialog, reads a CSV, and populates the m/z text box."""
		import pandas as pd
        # Open a "File Open" dialog window
		file_path = qt.QFileDialog.getOpenFileName(
            None, 
            "Select m/z CSV File", 
            "", 
            "CSV Files (*.csv)"
        )

        # If the user clicked Cancel, just stop
		if not file_path:
			return
		
		try:
            # Read the CSV
			df = pd.read_csv(file_path)

            # Column Detection
			target_col = None
			for col in df.columns:
				col_clean = str(col).strip().lower()
				if col_clean in ['m/z', 'mz', 'mass', 'm.z']:
					target_col = col
					break
            
            # Fallback: If no matching header is found, just grab the first column
			if target_col is None:
				target_col = df.columns[0]

            # Extract the numbers, drop empty rows, and convert to strings
			mz_values = df[target_col].dropna().astype(str).tolist()
            
            # Join them with commas
			mz_string = ", ".join(mz_values)

            # Inject the string directly into the existing text box
			self.ui.inputtedpeakslineedit.setText(mz_string)
            
            # Let the user know it worked
			slicer.util.infoDisplay(f"Successfully loaded {len(mz_values)} peaks from column: '{target_col}'", "CSV Loaded")

		except Exception as e:
			slicer.util.errorDisplay(f"Failed to load CSV file:\n{str(e)}")

	def onLinkClicked(self, url):
		"""Catches the clicked link and opens it inside the Red window."""
		self.internalBrowser.setUrl(url.toString())
		self.addressBar.setText(url.toString())
		slicer.app.layoutManager().setLayout(self.browserOnlyLayoutId)


	def hideInternalBrowser(self):
		"""Hides the web browser and restores the full Red slice view."""
		self.webContainer.hide()
		
		# Restore Slicer's native widgets
		self.redWidget.sliceController().show()
		self.redWidget.sliceView().show()
		
		self.internalBrowser.url = "about:blank"

	def onTableLinkClicked(self, link_string):
		"""Helper function to convert table string links into QUrls for the internal browser."""
		# Convert the string to a QUrl and pass it to your existing browser function
		url_object = qt.QUrl(link_string)
		self.onLinkClicked(url_object)

	# --- End of Robert Peak Labeling Tab Functions ---

	### Multi-slide alignment tab

	def onSelectToAlign(self):
		# gets the list of files they are trying to merge and shows them in the viewer
		fileExplorer = qt.QFileDialog()
		filePaths = fileExplorer.getOpenFileNames(None, "Open CSV datasets", "", "CSV Files (*.csv);;All Files (*)")
		for filePath in filePaths:
			if filePath not in self.files:
				i = len(self.files)
				self.files.add(filePath)
				self.ui.filesTable.setRowCount(len(self.files))
				delete_btn = qt.QPushButton('Delete')
				delete_btn.clicked.connect(self.createDeleteButtonClickedHandler(filePath))
				self.ui.filesTable.setItem(i, 0, qt.QTableWidgetItem(filePath))
				self.ui.filesTable.setCellWidget(i, 1, delete_btn)

	# Wrapper functions to let us pass the filenames to the button press handler
	def createDeleteButtonClickedHandler(self, path):
		def deleteButtonPressed(checked):
			self.handleDeleteButtonClicked(checked, path)
		return deleteButtonPressed

	def handleDeleteButtonClicked(self, checked, path=None):
		self.files.remove(path)
		r, numRows = 0, self.ui.filesTable.rowCount
		while r < numRows:
			if self.ui.filesTable.item(r, 0).text() == path:
				self.ui.filesTable.removeRow(r)
				self.ui.filesTable.setRowCount(len(self.files))
				numRows -= 1
			r += 1

	def onLoadToAlign(self):
		info, mz_range = self.logic.load_alignment_files(list(self.files))
		self.ui.fileInfoAlign.setText(info)

		self.ui.alignPreviewStart.setText(mz_range[0])
		self.ui.alignPreviewEnd.setText(mz_range[1])
	
	def onAlignPreview(self):
		params = {}
		params['mz_bandwidth'] = float(self.ui.KDEbandwidthVal.text)
		params['abundance_threshold'] = 1 - float(self.ui.sparsityVal.text)
		params['mz_resolution'] = params['mz_bandwidth']/2
		params['ion_count_method'] = self.ui.sparsityLevel.currentText
		params['preview'] = [float(self.ui.alignPreviewStart.text), float(self.ui.alignPreviewEnd.text)]
		params['savepath'] = None
		params['file_names'] = [os.path.splitext(os.path.basename(x))[0] for x in list(self.files)]

		self.logic.batch_peak_alignment(params)

	def onAlignMatchMethod(self, text):
		if text == "Cluster":
			self.ui.AlignTolLabel.setVisible(False)
			self.ui.AlignTolVal.setVisible(False)
			self.ui.AlignBinLab.setVisible(False)
			self.ui.AlignBinVal.setVisible(False)
		elif text == "Tolerance":
			self.ui.AlignTolLabel.setVisible(True)
			self.ui.AlignTolVal.setVisible(True)
			self.ui.AlignBinLab.setVisible(True)
			self.ui.AlignBinVal.setVisible(True)

	def onMerge(self):
		# Merge csv files added by the user
		fileExplorer = qt.QFileDialog()
		defaultSave = list(self.files)[-1][:-4]+'_aligned'
		savepath = fileExplorer.getSaveFileName(None, "Save aligned dataset", defaultSave, "CSV Files (*.csv);;All Files (*)")
		print('save path:',savepath)

		params = {}
		params['mz_bandwidth'] = float(self.ui.KDEbandwidthVal.text)
		params['abundance_threshold'] = 1 - float(self.ui.sparsityVal.text)
		params['mz_resolution'] = params['mz_bandwidth']/2
		params['ion_count_method'] = self.ui.sparsityLevel.currentText
		params['savepath'] = savepath
		params['preview'] = None
		params['matching_method'] = self.ui.AlignMatchMethod.currentText
		if params['matching_method'] == "Tolerance":
			params['matching_tol'] = float(self.ui.AlignTolVal.text)
			params['matching_bin'] = self.ui.AlignBinVal.currentText.lower()

		retstr = self.logic.batch_peak_alignment(params)
		self.ui.alignmentTextBrowser.setText(retstr)

	
	### Dataset post-processing
	def onCsvSelect(self):
		fileExplorer = qt.QFileDialog()
		csvFilename = fileExplorer.getOpenFileName(None, "Open CSV dataset", "", "CSV Files (*.csv);;All Files (*)")
		self.csvForProcess = csvFilename
		if csvFilename:
			self.ui.PostProcessinglineEdit.setText(csvFilename)
			self.ui.PostProcessinglineEdit.setToolTip(csvFilename)
			self.onCsvLoad()

	def onCsvLoad(self):
		csv_info = self.logic.CsvLoad(self.ui.PostProcessinglineEdit.text)
		if csv_info:
			self.ui.csvInfo.setText(csv_info)
			self.ui.refIoncomboBox.clear()
			all_mz = self.logic.getCsvMzList()
			for mz in all_mz:
				self.ui.refIoncomboBox.addItem(mz)
			threshold = self.logic.getTUSthreshold()
			self.ui.thresholdValue.setText(str(threshold))

	
	# def onNormalizationState(self):
	# 	if self.ui.normalizeCheckbox.isChecked():
	# 		self.ui.normalizeTICoption.setEnabled(True)
	# 		self.ui.refNorm.setEnabled(True)
	# 		self.onIonNorm()
	# 	else:
	# 		self.ui.normalizeTICoption.setEnabled(False)
	# 		self.ui.refNorm.setEnabled(False)
	# 		self.ui.refIoncomboBox.setEnabled(False)

	def onNormMethodChange(self, text):
		if text == "Reference ion":
			self.ui.refionLabel.setVisible(True)
			self.ui.refIoncomboBox.setVisible(True)
			self.ui.thresholdLabel.setVisible(False)
			self.ui.thresholdValue.setVisible(False)
		elif text == "Total signal current (TSC)":
			self.ui.refionLabel.setVisible(False)
			self.ui.refIoncomboBox.setVisible(False)
			self.ui.thresholdLabel.setVisible(True)
			self.ui.thresholdValue.setVisible(True)
		else:
			self.ui.refionLabel.setVisible(False)
			self.ui.refIoncomboBox.setVisible(False)
			self.ui.thresholdLabel.setVisible(False)
			self.ui.thresholdValue.setVisible(False)

	def onDepNormMethodChange(self, text):
		if text == "Reference ion":
			self.ui.depRefIonLab.setVisible(True)
			self.ui.depComboboxIon.setVisible(True)
			self.ui.depNormThreshLab.setVisible(False)
			self.ui.depNormThresh.setVisible(False)
		elif text == "Total signal current (TSC)":
			self.ui.depRefIonLab.setVisible(False)
			self.ui.depComboboxIon.setVisible(False)
			self.ui.depNormThreshLab.setVisible(True)
			self.ui.depNormThresh.setVisible(True)
		else:
			self.ui.depRefIonLab.setVisible(False)
			self.ui.depComboboxIon.setVisible(False)
			self.ui.depNormThreshLab.setVisible(False)
			self.ui.depNormThresh.setVisible(False)

	def onNormalizationState(self):
		if self.ui.normalizeCheckbox.isChecked():
			self.ui.normMethodLabel.setEnabled(True)
			self.ui.normMethodComboBox.setEnabled(True)
			self.ui.refionLabel.setEnabled(True)
			self.ui.refIoncomboBox.setEnabled(True)
			self.ui.thresholdLabel.setEnabled(True)
			self.ui.thresholdValue.setEnabled(True)
		else:
			self.ui.normMethodLabel.setEnabled(False)
			self.ui.normMethodComboBox.setEnabled(False)
			self.ui.refionLabel.setEnabled(False)
			self.ui.refIoncomboBox.setEnabled(False)
			self.ui.thresholdLabel.setEnabled(False)
			self.ui.thresholdValue.setEnabled(False)

	# def onIonNorm(self):
	# 	if self.ui.refNorm.isChecked():
	# 		self.ui.refIoncomboBox.setEnabled(True)
	# 	else:
	# 		self.ui.refIoncomboBox.setEnabled(False)


	def onFilterState(self):
		if self.ui.spectrumFiltercheckBox.isChecked():
			State = True
		else:
			State = False
		self.ui.spectrumUpperband.setEnabled(State)
		self.ui.spectrumlowerBand.setEnabled(State)
		self.ui.lowLabel.setEnabled(State)
		self.ui.upLabel.setEnabled(State)

	def onIntFilterState(self):
		if self.ui.lowIntFiltercheckBox.isChecked():
			State = True
		else:
			State = False
		self.ui.lowIntFilterVal.setEnabled(State)
		self.ui.lowIntFilterValLabel.setEnabled(State)
		self.ui.lowIntFilterMethod.setEnabled(State)
		self.ui.lowIntFilterMethodLabel.setEnabled(State)

	def onVarFilterState(self):
		if self.ui.lowVarFiltercheckBox.isChecked():
			State = True
		else:
			State = False
		self.ui.lowVarFilterVal.setEnabled(State)
		self.ui.lowVarFilterValLabel.setEnabled(State)
		self.ui.lowVarFilterMethod.setEnabled(State)
		self.ui.lowVarFilterMethodLabel.setEnabled(State)

	def onAggState(self):
		if self.ui.pixelaggcheckBox.isChecked():
			State = True
		else:
			State = False
		self.ui.patchWidth.setEnabled(State)
		self.ui.patchStride.setEnabled(State)
		self.ui.partialPatch.setEnabled(State)
		self.ui.aggMode.setEnabled(State)
		self.ui.wLabel.setEnabled(State)
		self.ui.sLabel.setEnabled(State)
		self.ui.pLabel.setEnabled(State)
		self.ui.mLabel.setEnabled(State)
	
	def onApplyProcessing(self):
		# get normalize parameters
		if self.ui.normalizeCheckbox.isChecked():
			spec_normalization = self.ui.normMethodComboBox.currentText
			if spec_normalization == "Reference ion":
				normalization_param = self.ui.refIoncomboBox.currentText
			elif spec_normalization == "Total signal current (TSC)":
				normalization_param = float(self.ui.thresholdValue.text)
			else:
				normalization_param = None
		else:
			spec_normalization = None
			normalization_param = None
		# if self.ui.normalizeCheckbox.isChecked():
		# 	if self.ui.normalizeTICoption.isChecked():
		# 		spec_normalization = 'tic'
		# 	else:
		# 		spec_normalization = self.ui.refIoncomboBox.currentText
		# else:
		# 	spec_normalization = None
	
		# get spectrum filtering
		if self.ui.spectrumFiltercheckBox.isChecked():
			subband_selection = (int(self.ui.spectrumlowerBand.text),int(self.ui.spectrumUpperband.text))
		else:
			subband_selection = None
   
		# ion filtering
		if self.ui.lowIntFiltercheckBox.isChecked():
			lowInt_val = ( self.ui.lowIntFilterMethod.currentText, int(self.ui.lowIntFilterVal.text) )
		else:
			lowInt_val = None
		if self.ui.lowVarFiltercheckBox.isChecked():
			lowVar_val = ( self.ui.lowVarFilterMethod.currentText, int(self.ui.lowVarFilterVal.text) )
		else:
			lowVar_val = None

		# get pixel aggregation
		if self.ui.pixelaggcheckBox.isChecked():
			pixel_aggregation = (int(self.ui.patchWidth.text), int(self.ui.patchStride.text), self.ui.aggMode.currentText, int(self.ui.partialPatch.text))
		else:
			pixel_aggregation = None

		fileExplorer = qt.QFileDialog()
		defaultSave = self.csvForProcess[:-4]+'_processed.csv'
		savepath = fileExplorer.getSaveFileName(None, "Save processed dataset", defaultSave, "CSV Files (*.csv);;All Files (*)")
		print(savepath)


		processed_csv_info = self.logic.dataset_post_processing(spec_normalization, 
														  normalization_param, 
														  subband_selection, 
														  lowInt_val, 
														  lowVar_val, 
														  pixel_aggregation, savepath)

		retstr = 'Dataset successfully processed! \n'
		retstr += f'Processed dataset:\t {savepath} \n'
		retstr += processed_csv_info
		self.ui.postCsvinfo.setText(retstr)


	### Statistical tab
	def onSelectStatData(self):
		fileExplorer = qt.QFileDialog()
		csvFilename = fileExplorer.getOpenFileName(None, "Open CSV dataset", "", "CSV Files (*.csv);;All Files (*)")
		if csvFilename:
			self.ui.fileStat.setText(csvFilename)
			self.ui.fileStat.setToolTip(csvFilename)

			csv_info = self.logic.CsvLoad(csvFilename)
			if csv_info:
				self.ui.infoStat.setText(csv_info)
				self.ui.statIonCombo.clear()
				all_mz = self.logic.getCsvMzList()
				for mz in all_mz:
					self.ui.statIonCombo.addItem(mz)

			classes = self.logic.classes
			unique_classes = np.unique(classes)

			self.ui.statClassConfig.clear()
			self.ui.statGroup1combo.clear()
			self.ui.statGroup2combo.clear()

			self.ui.statGroup1Lab.setVisible(False)
			self.ui.statGroup1combo.setVisible(False)
			self.ui.statGroup2Lab.setVisible(False)
			self.ui.statGroup2combo.setVisible(False)

			if len(unique_classes)==1:
				self.ui.statClassConfig.addItem("Single class")
			elif len(unique_classes)==2:
				self.ui.statClassConfig.addItem("All classes - binary")
			elif len(unique_classes)>2:
				self.ui.statClassConfig.addItem("All classes")
				self.ui.statClassConfig.addItem("Binary - one versus the rest")
				self.ui.statClassConfig.addItem("Binary - two classes")
				self.ui.statClassConfig.setCurrentText("All classes")

	def onStatConfigChange(self, text):
		if text in ["Single class", "All classes - binary", "All classes"]:
			self.ui.statGroup1Lab.setVisible(False)
			self.ui.statGroup1combo.setVisible(False)
			self.ui.statGroup2Lab.setVisible(False)
			self.ui.statGroup2combo.setVisible(False)
		elif text == "Binary - one versus the rest":
			classes = np.unique(self.logic.classes)
			self.ui.statGroup1combo.clear()
			for label in classes:
				self.ui.statGroup1combo.addItem(label)
			self.ui.statGroup1Lab.setVisible(True)
			self.ui.statGroup1combo.setVisible(True)
			self.ui.statGroup2Lab.setVisible(False)
			self.ui.statGroup2combo.setVisible(False)
		else:
			classes = np.unique(self.logic.classes)
			self.ui.statGroup1combo.clear()
			for label in classes:
				self.ui.statGroup1combo.addItem(label)
			self.ui.statGroup2combo.clear()
			for label in classes:
				self.ui.statGroup2combo.addItem(label)
			self.ui.statGroup1Lab.setVisible(True)
			self.ui.statGroup1combo.setVisible(True)
			self.ui.statGroup2Lab.setVisible(True)
			self.ui.statGroup2combo.setVisible(True)
	
	def onPlotDIstribution(self):
		self.logic.plot_latent_pca()

	def GetStatConfigParameters(self):
		label_1, label_2 = None, None
		if self.ui.statGroup2combo.isVisible():
			label_1 = self.ui.statGroup1combo.currentText
			label_2 = self.ui.statGroup2combo.currentText
		elif self.ui.statGroup1combo.isVisible():
			label_1 = self.ui.statGroup1combo.currentText

		return [label_1, label_2]

	def onBoxPlot(self):
		mz_ref = self.ui.statIonCombo.currentText
		label_config = self.GetStatConfigParameters()
		df_summary = self.logic.BoxPlot(mz_ref, label_config)
		slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUpYellowSliceView)
		slicer.util.resetSliceViews()
		
		# view the table below the ion images
		customLayoutId = 90
		customLayout = """
		<layout type="horizontal" split="true">
		<item>
			<view class="vtkMRMLSliceNode" singletontag="Yellow"/>
		</item>
		<item>
			<view class="vtkMRMLTableViewNode" singletontag="BoxPlotTable"/>
		</item>
		</layout>
		"""
		slicer.app.layoutManager().layoutLogic().GetLayoutNode().AddLayoutDescription(customLayoutId, customLayout)
		slicer.app.layoutManager().setLayout(customLayoutId)

		# create a table node
		tableNode = slicer.mrmlScene.GetFirstNodeByName("BoxplotStats")
		if not tableNode:
			tableNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLTableNode', 'BoxplotStats')
		else:
			tableNode.RemoveAllColumns()

		for col in df_summary.columns:
			array = vtk.vtkVariantArray()
			array.SetName(str(col))
			for val in df_summary[col]:
				array.InsertNextValue(vtk.vtkVariant(str(val)))
			tableNode.AddColumn(array)

		# lock the table
		tableNode.SetUseColumnTitleAsColumnHeader(True)
		tableNode.SetUseFirstColumnAsRowHeader(True)
		tableNode.SetLocked(True)

		# set the table view node
		tableViewNodes = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
		if tableViewNodes:
			tableViewNode = tableViewNodes[0]
			tableViewNode.SetTableNodeID(tableNode.GetID())

		# TableView node (singleton tag "BoxPlotTable")
		tableViewNode = slicer.mrmlScene.GetSingletonNode("BoxPlotTable", "vtkMRMLTableViewNode")
		# Fallback if needed:
		if not tableViewNode:
			tvs = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
			tableViewNode = tvs[0] if tvs else None

		if tableViewNode:
			tableViewNode.SetTableNodeID(tableNode.GetID())

	def onANOVA(self):
		self.onRunStat(test_method='anova', tabName = "ANOVA")
	
	def onTtest(self):
		self.onRunStat(test_method='ttest', tabName = "tTest")

	def onVolcano(self):
		self.onRunStat(test_method='ttest', tabName = "Volcano", return_volcano=True)

	def onRunStat(self, test_method, tabName, return_volcano=False):
		label_config = self.GetStatConfigParameters()

		if test_method == 'anova':
			stat_results = self.logic.runANOVA(label_config)
		elif test_method == 'ttest':
			stat_results = self.logic.runTtest(label_config, return_volcano)

		if stat_results is None:
			warning_dialog = qt.QMessageBox()
			warning_dialog.setIcon(qt.QMessageBox.Warning)
			warning_dialog.setText("This method is not compatible with non-binary data. Please choose a binary configuration.")
			warning_dialog.setWindowTitle("Non-binary warning")
			warning_dialog.setStandardButtons(qt.QMessageBox.Ok)
			warning_dialog.exec_()
			return False

		# set the layout
		if not return_volcano:
			customLayoutId = 91
			customLayout = """
			<layout type="horizontal" split="true">
			<item>
				<view class="vtkMRMLTableViewNode" singletontag="StatTable"/>
			</item>
			<item>
				<view class="vtkMRMLSliceNode" singletontag="Yellow"/>
			</item>
			</layout>
			"""
			YellowCompNode = slicer.util.getNode("vtkMRMLSliceCompositeNodeYellow")
			YellowCompNode.SetBackgroundVolumeID("")
			slicer.app.layoutManager().layoutLogic().GetLayoutNode().AddLayoutDescription(customLayoutId, customLayout)
			slicer.app.layoutManager().setLayout(customLayoutId)

		else:
			customLayoutId = 92
			customLayout = """
			<layout type="vertical" split="true">
			<item>
				<layout type="horizontal" split="true">
				<item>
					<view class="vtkMRMLTableViewNode" singletontag="StatTable"/>
				</item>
				<item>
					<view class="vtkMRMLSliceNode" singletontag="Yellow"/>
				</item>
				</layout>
			</item>
			<item>
				<view class="vtkMRMLSliceNode" singletontag="Red"/>
			</item>
			</layout>
			"""
			YellowCompNode = slicer.util.getNode("vtkMRMLSliceCompositeNodeYellow")
			YellowCompNode.SetBackgroundVolumeID("")
			slicer.app.layoutManager().layoutLogic().GetLayoutNode().AddLayoutDescription(customLayoutId, customLayout)
			slicer.app.layoutManager().setLayout(customLayoutId)
			slicer.util.resetSliceViews()

		# create a table node
		tableNode = slicer.mrmlScene.GetFirstNodeByName(tabName)
		if not tableNode:
			tableNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLTableNode', tabName)
		else:
			tableNode.RemoveAllColumns()

		for col in stat_results.columns:
			array = vtk.vtkVariantArray()
			array.SetName(str(col))
			for val in stat_results[col]:
				array.InsertNextValue(vtk.vtkVariant(str(val)))
			tableNode.AddColumn(array)

		# lock the table
		tableNode.SetUseColumnTitleAsColumnHeader(True)
		tableNode.SetLocked(True)

		# # set the table view node
		# tableViewNodes = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
		# if tableViewNodes:
		# 	tableViewNode = tableViewNodes[0]
		# 	tableViewNode.SetTableNodeID(tableNode.GetID())

		# TableView node (singleton tag "StatTable")
		tableViewNode = slicer.mrmlScene.GetSingletonNode("StatTable", "vtkMRMLTableViewNode")
		# Fallback if needed:
		if not tableViewNode:
			tvs = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
			tableViewNode = tvs[0] if tvs else None

		if tableViewNode:
			tableViewNode.SetTableNodeID(tableNode.GetID())

		# interactive boxplot on cell click
		self.tableNode = tableNode

		tableView = None
		for w in qt.QApplication.instance().allWidgets():
			if isinstance(w, slicer.qMRMLTableView) and w.mrmlTableNode() is tableNode:
				tableView = w
				break
		if not tableView:
			slicer.util.errorDisplay("Table view for 'Stat' is not open")
			raise RuntimeError("No open Table View for 'Stat'")
		
		try:
			tableView.clicked.disconnect()
		except Exception:
			pass
		tableView.clicked.connect(self.onStatCellClicked)

	def onStatCellClicked(self, index):
		row = index.row()
		mz_ref = self.tableNode.GetCellText(row, 0)
		label_config = self.GetStatConfigParameters()
		df_summary = self.logic.BoxPlot(mz_ref, label_config)
		yellowNode = slicer.util.getNode("vtkMRMLSliceNodeYellow")
		yellowLogic = slicer.app.applicationLogic().GetSliceLogic(yellowNode)
		yellowLogic.FitSliceToAll()


	### Model training tab
	def onRankMethodChange(self, text):
		if text == "Linear SVC":
			self.ui.FRankParamLab.setVisible(True)
			self.ui.FRankParamVal.setVisible(True)
			self.ui.FRankParamLab.setText("C")
			self.ui.FRankParamLab.setToolTip("Regularization strength")
			self.ui.FRankParamVal.setText("1.0")
		elif text == "PLS-DA":
			self.ui.FRankParamLab.setVisible(True)
			self.ui.FRankParamVal.setVisible(True)
			self.ui.FRankParamLab.setText("n_components")
			self.ui.FRankParamLab.setToolTip("Number of components")
			self.ui.FRankParamVal.setText("2")
		elif text == 'LDA':
			self.ui.FRankParamLab.setVisible(False)
			self.ui.FRankParamVal.setVisible(False)


	def onFeatureRank(self):
		rankMethod = self.ui.FRankMethod.currentText
		rankParam = float(self.ui.FRankParamVal.text)
		ranked_df = self.logic.feature_ranking(rankMethod, rankParam)

		customLayoutId = 70
		customLayout = """
		<layout type="vertical">
		<item>
			<view class="vtkMRMLTableViewNode" singletontag="FeatureTable"/>
		</item>
		</layout>
		"""
		slicer.app.layoutManager().layoutLogic().GetLayoutNode().AddLayoutDescription(customLayoutId, customLayout)
		slicer.app.layoutManager().setLayout(customLayoutId)

		## create a table node
		tableNode = slicer.mrmlScene.GetFirstNodeByName("Ranking")
		if not tableNode:
			tableNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLTableNode', 'Ranking')
		else:
			tableNode.RemoveAllColumns()

		for col in ranked_df.columns:
			array = vtk.vtkVariantArray()
			array.SetName(str(col))
			for val in ranked_df[col]:
				array.InsertNextValue(vtk.vtkVariant(str(val)))
			tableNode.AddColumn(array)


		# TableView node (singleton tag "FeatureTable")
		tableViewNode = slicer.mrmlScene.GetSingletonNode("FeatureTable", "vtkMRMLTableViewNode")
		# Fallback if needed:
		if not tableViewNode:
			tvs = slicer.util.getNodesByClass("vtkMRMLTableViewNode")
			tableViewNode = tvs[0] if tvs else None

		if tableViewNode:
			tableViewNode.SetTableNodeID(tableNode.GetID())

		## lock the table
		tableNode.SetUseColumnTitleAsColumnHeader(True)
		tableNode.SetLocked(True)

	def onSelMethodChange(self, text):
		if text == "Top ranked":
			self.ui.FnumberLabel.setVisible(True)
			self.ui.FnumberValue.setVisible(True)
			self.ui.FSelManualUpload.setVisible(False)
		elif text == "Manual":
			self.ui.FnumberLabel.setVisible(False)
			self.ui.FnumberValue.setVisible(False)
			self.ui.FSelManualUpload.setVisible(True)
		elif text == "None":
			self.ui.FnumberLabel.setVisible(False)
			self.ui.FnumberValue.setVisible(False)
			self.ui.FSelManualUpload.setVisible(False)

	def onFeatureListUpload(self):
		fileExplorer = qt.QFileDialog()
		featureFiles = fileExplorer.getOpenFileName(None, "Upload feature indices", "", "CSV Files (*.csv);;All Files (*)")
		if featureFiles:
			df = pd.read_csv(featureFiles, header=None)
			self.logic.manual_features_indices = [int(x) for x in df.values.ravel()]
			print("Manual feature load compeleted")
			print(self.logic.manual_features_indices)

	def onSelectModelData(self):
		fileExplorer = qt.QFileDialog()
		csvFilename = fileExplorer.getOpenFileName(None, "Open CSV dataset", "", "CSV Files (*.csv);;All Files (*)")
		if csvFilename:
			self.ui.ModellineEdit.setText(csvFilename)
			self.ui.ModellineEdit.setToolTip(csvFilename)
			self.onModellingLoad()

	def onModellingLoad(self):
		# loads the file for modelling and does confirmation pop up 
		dataset_info = self.logic.modellingFileLoad(self.ui.ModellineEdit.text)
		if not dataset_info:
			dataset_info = 'Error in file load. Please check the console error and try again.'
		else:
			self.ui.dataInformation_2.setText(dataset_info)

		# from functools import partial
		self.cases_config = {}
		names_list = self.logic.custom_split()
		self.ui.namesTable.setRowCount(len(names_list))
		self.ui.namesTable.setColumnCount(4)
		self.ui.namesTable.setHorizontalHeaderLabels(['Name', 'Train', 'Test', 'Validation'])

		for i, name in enumerate(names_list):
			train_btn, test_btn, val_btn = qt.QCheckBox(), qt.QCheckBox(), qt.QCheckBox()
			train_btn.clicked.connect(self.createRadioButtonClickedHandler(name, train_btn))
			test_btn.clicked.connect(self.createRadioButtonClickedHandler(name, test_btn))
			val_btn.clicked.connect(self.createRadioButtonClickedHandler(name, val_btn))
			self.cases_config[name] = [train_btn, test_btn, val_btn]
			self.ui.namesTable.setItem(i, 0, qt.QTableWidgetItem(name))
			self.ui.namesTable.setCellWidget(i, 1, train_btn)
			self.ui.namesTable.setCellWidget(i, 2, test_btn)
			self.ui.namesTable.setCellWidget(i, 3, val_btn)
			train_btn.setChecked(True)

		self.ui.namesTable.setColumnHidden(3, True)

		if self.ui.randomSplit.isChecked():
			self.logic.set_split('random')
		elif self.ui.allTrain.isChecked():
			self.logic.set_split('all_train')
		elif self.ui.XVall.isChecked():
			self.logic.set_split('cross_val')
		

	
	def updateCustomConfig(self):
		"""
  		Let logic keep track of what cases are train/test.
		"""
		train_cases, test_cases, val_cases = set(), set(), set()
		for case, buttons in self.cases_config.items():
			train_btn, test_btn, val_btn = buttons
			if train_btn.isChecked(): train_cases.add(case)
			if test_btn.isChecked(): test_cases.add(case)
			if val_btn.isChecked(): val_cases.add(case)
		self.logic.update_test_cases(test_cases)
		self.logic.update_train_cases(train_cases)
		self.logic.update_val_cases(val_cases)
	

	def onRandomSplit(self):
		"""
  		Let logic keep track of what split is being used.
		"""
		self.ui.namesTable.hide()
		self.ui.namesTableLabel.hide()
		self.logic.set_split('random')
		self.ui.saveModelcheckBox.setVisible(True)

	def onAllTrainSplit(self):
		"""
  		Let logic keep track of what split is being used.
		"""
		self.ui.namesTable.hide()
		self.ui.namesTableLabel.hide()
		self.logic.set_split('all_train')
		self.ui.saveModelcheckBox.setVisible(True)

	def onCustomSplit(self):
		"""
		Populate namesTable.
  		
		For each case, adds a row with the case name and checkboxes
		so the user can select if they want it to be in train/test.
		"""
		self.ui.namesTable.show()
		self.ui.namesTableLabel.show()
		self.logic.set_split('custom')
		self.ui.saveModelcheckBox.setVisible(True)
			
	def onCrossVal(self):
		"""
  		Let logic keep track of what split is being used.
		"""
		self.ui.namesTable.hide()
		self.ui.namesTableLabel.hide()
		self.logic.set_split('cross_val')
		self.ui.saveModelcheckBox.setChecked(False)
		self.ui.saveModelcheckBox.setVisible(False)
	# Wrapper functions to let us pass the buttons to the function,
	# making the checkboxes exclusive. 

	def createRadioButtonClickedHandler(self, name, checkbox):
		def radioButtonClicked(checked):
			self.handleRadioButtonClicked(checked, name, checkbox)
		return radioButtonClicked

	def handleRadioButtonClicked(self, checked, name=None, checkbox=None):
		train_btn, test_btn, val_btn = self.cases_config[name]
		if (train_btn.isChecked() or val_btn.isChecked()) and checkbox == test_btn and checked:
			train_btn.setChecked(False)
			val_btn.setChecked(False)
			test_btn.setChecked(True)
		if (test_btn.isChecked() or val_btn.isChecked()) and checkbox == train_btn and checked:
			test_btn.setChecked(False)
			val_btn.setChecked(False)
			train_btn.setChecked(True)
		if (test_btn.isChecked() or train_btn.isChecked()) and checkbox == val_btn and checked:
			test_btn.setChecked(False)
			train_btn.setChecked(False)
			val_btn.setChecked(True)
   
	def onModelTrain(self):
		"""
		Train model according to user specifications.
		"""
		savepath = None
		if self.ui.saveModelcheckBox.isChecked():
			fileExplorer = qt.QFileDialog()
			defaultSave = self.logic.modellingFile[:-4]+"_model.pkl"
			savepath = fileExplorer.getSaveFileName(None, "Save model pipeline", defaultSave, "Pickle Files (*.pkl);;All Files (*)")
			print(savepath)

		self.updateCustomConfig()
		
		self.logic.model_type = self.ui.ModelSelectCombobox.currentText
		self.logic.train_balancing = self.ui.BalanceComBox.currentText

		self.logic.model_param1 = float(self.ui.MLparam1.text)
		self.logic.model_param2 = float(self.ui.MLparam2.text)

		## feature selection
		feature_select_method = self.ui.FSelMethod.currentText
		if feature_select_method == "None":
			self.logic.selected_features_indices = None
		elif feature_select_method == "Top ranked":
			n_features = int(self.ui.FnumberValue.text)
			self.logic.selected_features_indices = self.logic.ranked_features_indices[:n_features]
		elif feature_select_method == "Manual":
			self.logic.selected_features_indices = self.logic.manual_features_indices
		
		accuracystring = self.logic.runModel(savepath)
		if not accuracystring:
			self.ui.textBrowser.setText('An error occured. Please check the console for details.')
		else:
			self.ui.textBrowser.setText(accuracystring)
			self.model_results = accuracystring
			self.ui.tabWidget.setCurrentIndex(8)

  
	def onMLMethod(self, text):
		if text=="PCA-LDA":
			self.ui.MLlabel1.setText("n_components")
			self.ui.MLlabel1.setToolTip("Number of components, or variance explained")
			self.ui.MLparam1.setText("0.99")

			self.ui.MLlabel2.setVisible(False)
			self.ui.MLparam2.setVisible(False)
		
		elif text=="Linear SVC":
			self.ui.MLlabel1.setText("C")
			self.ui.MLlabel1.setToolTip("Regularization strength")
			self.ui.MLparam1.setText("1.0")

			self.ui.MLlabel2.setVisible(False)
			self.ui.MLparam2.setVisible(False)

		elif text=="Random Forest":
			self.ui.MLlabel1.setText("n_estimators")
			self.ui.MLlabel1.setToolTip("Number of trees in the forest")
			self.ui.MLparam1.setText("100")

			self.ui.MLlabel2.setVisible(False)
			self.ui.MLparam2.setVisible(False)

		elif text=="PLS-DA":
			self.ui.MLlabel1.setText("n_components")
			self.ui.MLlabel1.setToolTip("Number of components")
			self.ui.MLparam1.setText("2")

			self.ui.MLlabel2.setVisible(False)
			self.ui.MLparam2.setVisible(False)
			# self.ui.NLVisLabel2.setText("min_dist")
			# self.ui.NLVisLabel2.setToolTip("0.0 - 0.99")
			# self.ui.NLVisParam2.setText("0.1")
			


	### Model deployment tab
	def onDeploySelect(self):
		file_path = self.textFileSelect()
		if file_path:
			self.ui.deployLoclineEdit.setText(file_path)
			self.ui.deployLoclineEdit.setToolTip(file_path)
			self.onDeployLoad()
   
	def onDeployLoad(self):
		# on the text file load runs the text file load and shows the confirmation button
		self.logic.textFileLoad(self.ui.deployLoclineEdit.text)
		info = self.logic.getDataInformation()
		self.ui.deployInfo.setText(info)
		## make the visualization options available for this slide
		self.visRenormalize()
		self.logic.heatmap_display()
		self.populateMzLists()
		# self.updateDepVisList()
  
	def onDeployModelSel(self):
		fileExplorer = qt.QFileDialog()
		pklFilename = fileExplorer.getOpenFileName(None, "Open model pipeline", "", "Pickle Files (*.pkl);;All Files (*)")
		self.pklFilename = pklFilename
		if pklFilename:
			self.ui.deployModellineEdit.setText(pklFilename)
			self.ui.deployModellineEdit.setToolTip(pklFilename)
			self.onDeployModelLoad()

	def onDeployModelLoad(self):
		pipelineInfo = self.logic.loadPipeline(self.ui.deployModellineEdit.text)
		self.ui.deployModelInfo.setText(pipelineInfo)
		self.ui.depComboboxIon.clear()
		for mz in self.logic.DmzRef:
			self.ui.depComboboxIon.addItem(mz)

	# def onDeployNormCheck(self):
	# 	if self.ui.deployNormcheck.isChecked():
	# 		self.ui.depRadioTIC.setEnabled(True)
	# 		self.ui.depRadioIon.setEnabled(True)
	# 		self.onDepNormRadioToggle()
	# 	else:
	# 		self.ui.depRadioTIC.setEnabled(False)
	# 		self.ui.depRadioIon.setEnabled(False)
	# 		self.ui.depComboboxIon.setEnabled(False)

	def onDeployNormCheck(self):
		if self.ui.deployNormcheck.isChecked():
			self.ui.depNormMethLab.setEnabled(True)
			self.ui.depNormMethod.setEnabled(True)
			self.ui.depRefIonLab.setEnabled(True)
			self.ui.depComboboxIon.setEnabled(True)
			self.ui.depNormThreshLab.setEnabled(True)
			self.ui.depNormThresh.setEnabled(True)
		else:
			self.ui.depNormMethLab.setEnabled(False)
			self.ui.depNormMethod.setEnabled(False)
			self.ui.depRefIonLab.setEnabled(False)
			self.ui.depComboboxIon.setEnabled(False)
			self.ui.depNormThreshLab.setEnabled(False)
			self.ui.depNormThresh.setEnabled(False)

	def onDepNormRadioToggle(self):
		if self.ui.depRadioIon.isChecked():
			self.ui.depComboboxIon.setEnabled(True)
		else:
			self.ui.depComboboxIon.setEnabled(False)

	def onDeployAggCheck(self):
		if self.ui.deployAGGcheck.isChecked():
			State = True
		else:
			State = False
		self.ui.aggwLabel.setEnabled(State)
		self.ui.aggmodeLabel.setEnabled(State)
		self.ui.aggW.setEnabled(State)
		self.ui.aggMode_2.setEnabled(State)
		
	def onDepMaskcheck(self):
		if self.ui.depMaskcheck.isChecked():
			state = True
		else:
			state = False
		self.ui.depGoVisButton.setEnabled(state)
		self.ui.depVisSelLabel.setEnabled(state)
		self.ui.depVisCombo.setEnabled(state)
		self.ui.depPCAVis.setEnabled(state)
		self.ui.depGoSegEdButton.setEnabled(state)
		self.ui.depSegListLabel.setEnabled(state)
		self.ui.depSegListCombo.setEnabled(state)

	def onDepGoVis(self):
		self.ui.tabWidget.setCurrentIndex(2)

	def onDepGoSeg(self):
		sourceVolumeNode = self.ui.depVisCombo.currentNode()
		slicer.util.selectModule("SegmentEditor")

		# set master volume and geometry
		segmentEditorNode = slicer.util.getNodesByClass('vtkMRMLSegmentEditorNode')[0]
		segmentationNode = slicer.util.getNodesByClass('vtkMRMLSegmentationNode')[0]
		segmentationNode.SetReferenceImageGeometryParameterFromVolumeNode(sourceVolumeNode)
		segmentEditorWidget = slicer.qMRMLSegmentEditorWidget()
		segmentEditorWidget.setMRMLScene(slicer.mrmlScene)
		segmentEditorWidget.setMRMLSegmentEditorNode(segmentEditorNode)
		segmentEditorWidget.setSegmentationNode(segmentationNode)
		segmentEditorWidget.setSourceVolumeNode(sourceVolumeNode)

		# set the display view and link views
		RedCompNode = slicer.util.getNode("vtkMRMLSliceCompositeNodeRed")
		RedCompNode.SetBackgroundVolumeID(sourceVolumeNode.GetID())
		RedNode = slicer.util.getNode("vtkMRMLSliceNodeRed")
		RedNode.SetOrientation("Axial")

		slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUpRedSliceView)
		slicer.util.resetSliceViews()

	# def onDepSegListUpdate(self):
	# 	segmentationNode = slicer.util.getNodesByClass('vtkMRMLSegmentationNode')[0]
	# 	segmentation = segmentationNode.GetSegmentation()
	# 	segIDs = segmentation.GetSegmentIDs()
	# 	segNames = [segmentation.GetSegment(segID).GetName() for segID in segIDs]

	# 	self.ui.depSegListCombo.clear()
	# 	self.ui.depSegListCombo.addItem('None')
	# 	for segName in segNames:
	# 		self.ui.segVollist1.addItem(segName)

	def onApplyDeployment(self):
		# spectrum normalization
		if self.ui.deployNormcheck.isChecked():
			spec_normalization = self.ui.depNormMethod.currentText
			if spec_normalization == "Reference ion":
				normalization_param = float(self.ui.depComboboxIon.currentText)
			elif spec_normalization == "Total signal current (TSC)":
				normalization_param = float(self.ui.depNormThresh.text)
			else:
				normalization_param = None
		else:
			spec_normalization = None
			normalization_param = None

		# if self.ui.deployNormcheck.isChecked():
		# 	if self.ui.depRadioTIC.isChecked():
		# 		spec_normalization = 'tic'
		# 	else:
		# 		spec_normalization = self.ui.depComboboxIon.currentText
		# else:
		# 	spec_normalization = None
	
		# spectrum aggregation
		if self.ui.deployAGGcheck.isChecked():
			pixel_aggregation = ( int(self.ui.aggW.text), self.ui.aggMode_2.currentText )
		else:
			pixel_aggregation = None
		
		dep_mask = None
		if self.ui.depMaskcheck.isChecked():
			dep_mask = self.ui.depSegListCombo.currentText

		self.logic.model_deployment(spec_normalization, normalization_param, pixel_aggregation, dep_mask)

	### Boilerplate functions from template
			
	def cleanup(self):
		"""
		Called when the application closes and the module widget is destroyed.
		"""
		self.removeObservers()

	def enter(self):
		"""
		Called each time the user opens this module.
		"""
		# Make sure parameter node exists and observed
		self.initializeParameterNode()

	def exit(self):
		"""
		Called each time the user opens a different module.
		"""
		# Do not react to parameter node changes (GUI wlil be updated when the user enters into the module)
		self.removeObserver(self._parameterNode, vtk.vtkCommand.ModifiedEvent, self.updateGUIFromParameterNode)

	def onSceneStartClose(self, caller, event):
		"""
		Called just before the scene is closed.
		"""
		# Parameter node will be reset, do not use it anymore
		self.setParameterNode(None)

	def onSceneEndClose(self, caller, event):
		"""
		Called just after the scene is closed.
		"""
		# If this module is shown while the scene is closed then recreate a new parameter node immediately
		if self.parent.isEntered:
			self.initializeParameterNode()

	def initializeParameterNode(self):
		"""
		Ensure parameter node exists and observed.
		"""
		# Parameter node stores all user choices in parameter values, node selections, etc.
		# so that when the scene is saved and reloaded, these settings are restored.

		self.setParameterNode(self.logic.getParameterNode())

		# Select default input nodes if nothing is selected yet to save a few clicks for the user
		if not self._parameterNode.GetNodeReference("InputVolume"):
			firstVolumeNode = slicer.mrmlScene.GetFirstNodeByClass("vtkMRMLScalarVolumeNode")
			if firstVolumeNode:
				self._parameterNode.SetNodeReferenceID("InputVolume", firstVolumeNode.GetID())

	def setParameterNode(self, inputParameterNode):
		"""
		Set and observe parameter node.
		Observation is needed because when the parameter node is changed then the GUI must be updated immediately.
		"""

		if inputParameterNode:
			self.logic.setDefaultParameters(inputParameterNode)

		# Unobserve previously selected parameter node and add an observer to the newly selected.
		# Changes of parameter node are observed so that whenever parameters are changed by a script or any other module
		# those are reflected immediately in the GUI.
		if self._parameterNode:
			self.removeObserver(self._parameterNode, vtk.vtkCommand.ModifiedEvent, self.updateGUIFromParameterNode)
		self._parameterNode = inputParameterNode
		if self._parameterNode:
			self.addObserver(self._parameterNode, vtk.vtkCommand.ModifiedEvent, self.updateGUIFromParameterNode)

		# Initial GUI update
		self.updateGUIFromParameterNode()

	def updateGUIFromParameterNode(self, caller=None, event=None):
		"""
		This method is called whenever parameter node is changed.
		The module GUI is updated to show the current state of the parameter node.
		"""

		if not self._parameterNode or self._updatingGUIFromParameterNode:
			return

		# Make sure GUI changes do not call updateParameterNodeFromGUI (it could cause infinite loop)
		self._updatingGUIFromParameterNode = True

		# All the GUI updates are done
		self._updatingGUIFromParameterNode = False

	def updateParameterNodeFromGUI(self, caller=None, event=None):
		"""
		This method is called when the user makes any change in the GUI.
		The changes are saved into the parameter node (so that they are restored when the scene is saved and loaded).
		"""

		if not self._parameterNode or self._updatingGUIFromParameterNode:
			return

		wasModified = self._parameterNode.StartModify()  # Modify all properties in a single batch

		self._parameterNode.SetNodeReferenceID("InputVolume", self.ui.inputSelector.currentNodeID)
		self._parameterNode.SetNodeReferenceID("OutputVolume", self.ui.outputSelector.currentNodeID)
		self._parameterNode.SetParameter("Threshold", str(self.ui.imageThresholdSliderWidget.value))
		self._parameterNode.SetParameter("Invert", "true" if self.ui.invertOutputCheckBox.checked else "false")
		self._parameterNode.SetNodeReferenceID("OutputVolumeInverse", self.ui.invertedOutputSelector.currentNodeID)

		self._parameterNode.EndModify(wasModified)

	# def onApplyButton(self):
	# 	"""
	# 	Run processing when user clicks "Apply" button.
	# 	"""
	# 	try:
	# 		# Compute output
	# 		self.logic.process(self.ui.inputSelector.currentNode(), 
	# 											 self.ui.outputSelector.currentNode(),
	# 											 self.ui.imageThresholdSliderWidget.value, 
	# 											 self.ui.invertOutputCheckBox.checked)
	# 		# Compute inverted output (if needed)
	# 		if self.ui.invertedOutputSelector.currentNode():
	# 			# If additional output volume is selected then result with inverted threshold is written there
	# 			self.logic.process(self.ui.inputSelector.currentNode(), 
	# 												 self.ui.invertedOutputSelector.currentNode(),
	# 												 self.ui.imageThresholdSliderWidget.value, 
	# 												 not self.ui.invertOutputCheckBox.checked, 
	# 												 showResult=False)
	# 	except Exception as e:
	# 		slicer.util.errorDisplay("Failed to compute results: "+str(e))
	# 		import traceback
	# 		traceback.print_exc()

	def onReload(self):
		"""
		Modified the classic reload button since splitting the module
		into files requires a custom function. Code is taken from:
		https://discourse.slicer.org/t/python-scripted-module-development-reload-feature-for-multiple-files/6363/4 
		"""
		# logging.debug("Reloading MassVision")
		# packageName='MassVisionLib'
		# submoduleNames=['Logic', 'Utils']
		# import imp
		# f, filename, description = imp.find_module(packageName)
		# package = imp.load_module(packageName, f, filename, description)
		# for submoduleName in submoduleNames:
		# 	f, filename, description = imp.find_module(submoduleName, package.__path__)
		# 	try:
		# 		imp.load_module(packageName+'.'+submoduleName, f, filename, description)
		# 	finally:
		# 		f.close()
		# ScriptedLoadableModuleWidget.onReload(self)

		import importlib
		logging.debug("Reloading MassVision")
		package_name = 'MassVisionLib'
		submodules   = ['Logic', 'Utils']

		pkg = importlib.import_module(package_name)
		importlib.reload(pkg)

		for sub in submodules:
			full_name = f"{package_name}.{sub}"
			# ensure it’s imported so reload() can find it
			mod = importlib.import_module(full_name)
			importlib.reload(mod)

		ScriptedLoadableModuleWidget.onReload(self)


class SimHeatmapThresholdOverlay:
	def __init__(self, slider, refVolume):
		self.refVolume = refVolume
		self.sim = slicer.util.arrayFromVolume(self.refVolume)
		self.slider = slider

		labelNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLLabelMapVolumeNode', 'SimMask')
		m = vtk.vtkMatrix4x4()
		refVolume.GetIJKToRASMatrix(m)
		labelNode.SetIJKToRASMatrix(m)
		labelNode.SetOrigin(self.refVolume.GetOrigin())
		labelNode.SetSpacing(self.refVolume.GetSpacing())
		labelNode.SetAndObserveTransformNodeID(self.refVolume.GetTransformNodeID())
		self.labelNode = labelNode

		slicer.util.updateVolumeFromArray(self.labelNode, np.zeros_like(self.sim, dtype=np.uint8))
		self.labelNode.CreateDefaultDisplayNodes()

		### change mask color
		colorTableNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLColorTableNode", "SimHeatmapMaskColors")
		colorTableNode.SetTypeToUser()
		colorTableNode.SetNumberOfColors(2)
		colorTableNode.SetColor(0, "Background", 0.0, 0.0, 0.0, 0.0)
		colorTableNode.SetColor(1, "Mask", 0.5, 0.68, 0.5, 1.0)
		displayNode = self.labelNode.GetDisplayNode()
		displayNode.SetAndObserveColorNodeID(colorTableNode.GetID())

		lm = slicer.app.layoutManager()
		layoutNode = lm.layoutLogic().GetLayoutNode()
		twoUpId = 902
		twoUpXML = """
		<layout type="horizontal">
			<item><view class="vtkMRMLSliceNode" singletontag="Red"/></item>
			<item><view class="vtkMRMLSliceNode" singletontag="Green"/></item>
		</layout>
		"""
		layoutNode.AddLayoutDescription(twoUpId, twoUpXML)
		lm.setLayout(twoUpId)

		greenComp = lm.sliceWidget('Green').mrmlSliceCompositeNode()
		greenComp.SetBackgroundVolumeID(None)
		greenComp.SetLabelVolumeID(self.labelNode.GetID())
		greenComp.SetLabelOpacity(0.9)

		GreenNode = slicer.util.getNode("vtkMRMLSliceNodeGreen")
		GreenNode.SetOrientation("Axial")
		RedNode = slicer.util.getNode("vtkMRMLSliceNodeRed")

		markupNodes = slicer.mrmlScene.GetNodesByClass("vtkMRMLMarkupsNode")
		for markupNode in markupNodes:
			displayNode = markupNode.GetDisplayNode()
			displayNode.SetViewNodeIDs([RedNode.GetID()])

		redComp = lm.sliceWidget('Red').mrmlSliceCompositeNode()
		if redComp.GetLabelVolumeID() == self.labelNode.GetID():
			redComp.SetLabelVolumeID(None)

		lm.sliceWidget('Green').sliceLogic().FitSliceToAll()

		try:
			self.slider.valueChanged.disconnect()
		except Exception:
			pass
		self.slider.valueChanged.connect(self.onThresholdChanged)

		self.onThresholdChanged(self.slider.value)

	def onThresholdChanged(self, threshold):
		mask = (self.sim >= threshold).astype(np.uint8)
		slicer.util.updateVolumeFromArray(self.labelNode, mask)

		slicer.app.layoutManager().sliceWidget('Green').sliceLogic().FitSliceToAll()



###### Helper function for EmbedVision mode change
def recolorQIcon(icon, color="#80350E", size=qt.QSize(16, 16)):
    src = icon.pixmap(size)
    src = qt.QPixmap(src)  # copy

    targetColor = qt.QColor(color)

    # Start fully transparent
    colored = qt.QPixmap(src.size())
    colored.fill(qt.Qt.transparent)

    painter = qt.QPainter(colored)

    # 1) Copy original (with alpha) into `colored`
    painter.setCompositionMode(qt.QPainter.CompositionMode_Source)
    painter.drawPixmap(0, 0, src)

    # 2) Replace RGB with targetColor, keep alpha
    painter.setCompositionMode(qt.QPainter.CompositionMode_SourceIn)
    painter.fillRect(colored.rect(), targetColor)

    painter.end()

    return qt.QIcon(colored)

def recolorButtonIcon(btn, color="#80350E"):
    icon = btn.icon
    size = btn.iconSize
    newIcon = recolorQIcon(icon, color=color, size=size)
    btn.setIcon(newIcon)

def recolorTabIcon(tabWidget, index, color="#80350E"):
    icon = tabWidget.tabIcon(index)  
    size = tabWidget.tabBar().iconSize
    newIcon = recolorQIcon(icon, color=color, size=size)
    tabWidget.setTabIcon(index, newIcon)


def updateUITexts(ui):
	## change the labels for EmbedVision
	ui.label_importMSI.setText("Import Channel-rich Data")
	ui.label_importPATH.setText("Import Image")
	ui.label_ionNorm.setText("feature")
	ui.label_pixelNorm.setText("pixel")

	repl_map = {
		"ion": "feature",
		"ions": "features",
		"pixel spectrum": "pixel data",
		"spectrum": "pixel",
		"spectra": "pixels",
		"m/z": "",
	}

	# Build one big regex that matches ANY key in repl_map
	# Longer keys first so "pixel spectrum" wins over "spectrum"
	keys = sorted(repl_map.keys(), key=len, reverse=True)

	parts = []
	for key in keys:
		# If it's purely letters/spaces, treat as word/phrase → add \b boundaries
		if key.replace(" ", "").isalpha():
			part = r"\b" + re.escape(key) + r"\b"
		else:
			# e.g. "m/z" → just escape, no word boundaries
			part = re.escape(key)
		parts.append(part)

	pattern = re.compile("(" + "|".join(parts) + ")", re.IGNORECASE)

	def match_case(src: str, dst: str) -> str:
		"""Adjust dst to roughly match the casing style of src."""
		if not dst:
			return dst
		if src.isupper():
			return dst.upper()
		if src.istitle():
			# "Pixel Spectrum" → "Pixel Data"
			return dst.title()
		if src[0].isupper():
			# "Ion" → "Feature"
			return dst.capitalize()
		# default: all lower
		return dst.lower()

	def replacer(m: re.Match) -> str:
		src = m.group(0)          # actual matched text from UI
		key = src.lower()         # normalize for lookup
		base = repl_map.get(key, src)
		return match_case(src, base)

	for name, w in vars(ui).items():
		if isinstance(w, (qt.QPushButton, qt.QLabel)):
			old_text = w.text
			if not old_text:
				continue

			new_text = pattern.sub(replacer, old_text)

			if new_text != old_text:
				w.text = new_text
				# print(f"{name}: '{old_text}' -> '{new_text}'")

